"""
Sincronización bidireccional SQLite <-> Excel para revisión manual.

Flujo:
  1. export_to_excel()  → Genera Excel con detalle + catálogo + dropdowns
  2. Usuario revisa/corrige en Excel
  3. import_from_excel() → Detecta cambios, actualiza DB, marca MANUAL
  4. retrain_if_changed() → Reentrena modelo ML con correcciones

El Excel tiene:
  - Hoja "detalle": filas editables con dropdowns de categoría
  - Hoja "catalogo": referencia de combinaciones válidas
  - Hoja "documentos": datos de documentos (solo lectura)
  - Hoja "stock": resumen de inventario actual por código
  - Hoja "entradas": historial de entradas de compra para gestión de precios
  - Las filas SIN_CLASIFICAR y needs_review=1 se resaltan en amarillo/naranja
"""

from __future__ import annotations

import os
import sqlite3
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ─── Colores ───
_HEADER_FILL   = PatternFill("solid", fgColor="2B4C7E")
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
_REVIEW_FILL   = PatternFill("solid", fgColor="FFF3CD")   # amarillo: needs_review
_SIN_CLAS_FILL = PatternFill("solid", fgColor="F8D7DA")   # rojo claro: SIN_CLASIFICAR
_MANUAL_FILL   = PatternFill("solid", fgColor="D4EDDA")   # verde: corregido manual
_LOCKED_FONT   = Font(color="888888", italic=True)
_HYPERLINK_FONT = Font(color="0563C1", underline="single")
_THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Columnas de detalle que se exportan (orden importa)
_DETALLE_COLS = [
    "id_det", "id_doc", "linea", "codigo", "descripcion", "unidad", "cantidad",
    "precio_unitario", "monto_item",
    "razon_social", "giro", "fecha_emision",
    "categoria", "subcategoria", "tipo_gasto",
    "catalogo_costo_id", "confianza_categoria", "needs_review",
    "origen_clasificacion", "motivo_clasificacion",
]

# Columnas editables por el usuario (el resto es solo lectura visual)
_EDITABLE_COLS = {"categoria", "subcategoria", "tipo_gasto"}
_CATEGORY_COLS = ("categoria", "subcategoria", "tipo_gasto")
_NON_CATEGORY_SYNC_COLS = (
    "codigo",
    "descripcion",
    "unidad",
    "cantidad",
    "precio_unitario",
    "monto_item",
    "razon_social",
    "giro",
    "fecha_emision",
)
_NON_CATEGORY_NUMERIC_COLS = {"cantidad", "precio_unitario", "monto_item"}

# Columna oculta para detectar cambios
_HASH_COL = "hash_original"


def _to_excel_file_uri(path_value: str) -> Optional[str]:
    """
    Convierte una ruta local a URI file:// para hipervínculos en Excel.
    Priorizando la carpeta PDF configurada en el equipo actual.
    """
    raw = (path_value or "").strip()
    if not raw:
        return None

    base_pdf_raw = (os.getenv("RUTA_PDF_DTE_RECIBIDOS") or "").strip()
    base_pdf = Path(base_pdf_raw).expanduser() if base_pdf_raw else None

    p = Path(raw).expanduser()
    candidates: List[Path] = []

    # 1) Prioriza carpeta local configurada + nombre del archivo guardado en DB
    if base_pdf and p.name:
        candidates.append(base_pdf / p.name)

    # 2) Si DB tiene ruta relativa, también intentar dentro de la carpeta local
    if base_pdf and not p.is_absolute():
        candidates.append(base_pdf / p)

    # 3) Mantener compatibilidad con la ruta original guardada en DB
    candidates.append(p)

    chosen = None
    seen = set()
    for c in candidates:
        key = str(c).lower()
        if key in seen:
            continue
        seen.add(key)
        if c.exists():
            chosen = c
            break

    if chosen is None:
        chosen = candidates[0]

    try:
        return chosen.resolve().as_uri()
    except Exception:
        normalized = str(chosen).replace("\\", "/")
        if len(normalized) > 1 and normalized[1] == ":":
            return f"file:///{normalized}"
        return normalized


def _row_hash(cat: str, sub: str, tipo: str) -> str:
    """Hash MD5 corto de la clasificación original para detectar cambios."""
    raw = f"{(cat or '').strip().upper()}|{(sub or '').strip().upper()}|{(tipo or '').strip().upper()}"
    return hashlib.md5(raw.encode()).hexdigest()[:12]


def _coerce_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    # 1.234.567 o 1.234,50 -> formato CL habitual
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif re_fullmatch_miles(s):
        s = s.replace(".", "")
    try:
        return float(s)
    except Exception:
        return None


def re_fullmatch_miles(text: str) -> bool:
    # 1.234 o 12.345.678
    import re

    return bool(re.fullmatch(r"\d{1,3}(?:\.\d{3})+", text))


def _normalize_for_compare(col: str, value: Any) -> Any:
    if col in _NON_CATEGORY_NUMERIC_COLS:
        n = _coerce_number(value)
        return None if n is None else round(n, 6)
    return str(value or "").strip()


def _coerce_for_db(col: str, value: Any) -> Any:
    if col in _NON_CATEGORY_NUMERIC_COLS:
        n = _coerce_number(value)
        return float(n) if n is not None else 0.0
    return str(value or "").strip()


def _collect_non_category_changes(
    ws,
    col_map: Dict[str, int],
    db_by_id: Dict[str, sqlite3.Row],
    tracked_cols: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    if tracked_cols is None:
        tracked_cols = [c for c in _NON_CATEGORY_SYNC_COLS if c in col_map]
    out: List[Dict[str, Any]] = []
    if not tracked_cols:
        return out

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col_map["id_det"]]:
            continue
        id_det = str(row[col_map["id_det"]]).strip()
        db_row = db_by_id.get(id_det)
        if db_row is None:
            continue

        fields: Dict[str, Any] = {}
        changes_detail: List[Dict[str, Any]] = []
        for col in tracked_cols:
            old_raw = db_row[col] if col in db_row.keys() else None
            new_raw = row[col_map[col]]
            if _normalize_for_compare(col, old_raw) != _normalize_for_compare(col, new_raw):
                new_val = _coerce_for_db(col, new_raw)
                fields[col] = new_val
                changes_detail.append(
                    {
                        "columna": col,
                        "anterior": old_raw,
                        "nuevo": new_val,
                    }
                )

        if fields:
            out.append(
                {
                    "id_det": id_det,
                    "fields": fields,
                    "changes": changes_detail,
                }
            )
    return out


def preview_non_category_edits(
    excel_path: str,
    db_path: str,
    max_examples: int = 20,
) -> Dict[str, Any]:
    """
    Detecta cambios en columnas fuera de categoria/subcategoria/tipo_gasto.
    No aplica cambios; solo retorna resumen para confirmar con el usuario.
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    if "detalle" not in wb.sheetnames:
        wb.close()
        return {"ok": False, "error": "No se encontró la hoja 'detalle'"}

    ws = wb["detalle"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if "id_det" not in headers:
        wb.close()
        return {"ok": False, "error": "Columna 'id_det' no encontrada"}

    col_map = {h: i for i, h in enumerate(headers) if h}
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    try:
        detalle_cols = {r[1] for r in con.execute("PRAGMA table_info(detalle);").fetchall()}
        tracked_cols = [c for c in _NON_CATEGORY_SYNC_COLS if c in col_map and c in detalle_cols]
        cols = ["id_det"] + tracked_cols
        q = f"SELECT {', '.join(cols)} FROM detalle"
        db_rows = con.execute(q).fetchall()
        db_by_id = {str(r["id_det"]): r for r in db_rows}
        changes = _collect_non_category_changes(ws, col_map, db_by_id, tracked_cols=tracked_cols)
    finally:
        con.close()
        wb.close()

    n_fields = sum(len(c.get("fields", {})) for c in changes)
    return {
        "ok": True,
        "n_rows_with_changes": len(changes),
        "n_fields_changed": n_fields,
        "changes": changes,
        "examples": changes[: max(0, max_examples)],
    }


def _append_inventory_sheets(wb: Workbook, db_path: str) -> Tuple[int, int]:
    """
    Agrega hojas de inventario (`stock`, `entradas`) al workbook de exportación.
    """
    from app.core.DTE_Recibidos import inventory as INV

    stock_rows = INV.get_stock_summary(db_path=db_path, search_text="", limit=200000)
    entry_rows = INV.get_purchase_entries_for_export(db_path=db_path)

    # Hoja STOCK
    ws_stock = wb.create_sheet("stock")
    stock_headers = [
        "codigo",
        "descripcion_estandar",
        "unidad_base",
        "categoria",
        "tipo",
        "entradas",
        "salidas",
        "stock_actual",
        "ultima_fecha",
        "ultima_modificacion_tipo",
    ]
    for col_idx, header in enumerate(stock_headers, 1):
        cell = ws_stock.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER

    for row_idx, row in enumerate(stock_rows, 2):
        for col_idx, col_name in enumerate(stock_headers, 1):
            cell = ws_stock.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
            cell.border = _THIN_BORDER
    ws_stock.freeze_panes = "A2"

    # Hoja ENTRADAS
    ws_entries = wb.create_sheet("entradas")
    entries_headers = [
        "movimiento_id",
        "fecha",
        "codigo",
        "descripcion_estandar",
        "categoria",
        "tipo",
        "cantidad_entrada",
        "unidad_movimiento",
        "id_doc",
        "linea",
        "proveedor",
        "rut_emisor",
        "descripcion_dte_original",
        "cantidad_dte",
        "unidad_dte",
        "precio_unitario_dte",
        "monto_item_dte",
        "precio_unitario_calculado",
        "referencia_detalle",
    ]
    for col_idx, header in enumerate(entries_headers, 1):
        cell = ws_entries.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER

    for row_idx, row in enumerate(entry_rows, 2):
        for col_idx, col_name in enumerate(entries_headers, 1):
            cell = ws_entries.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
            cell.border = _THIN_BORDER
    ws_entries.freeze_panes = "A2"

    return len(stock_rows), len(entry_rows)


# =============================================================================
# EXPORTAR
# =============================================================================

def export_to_excel(
    db_path: str,
    output_path: str,
    only_pending: bool = False,
) -> Dict[str, Any]:
    """
    Exporta la DB a Excel con formato para revisión manual.

    Args:
        db_path: ruta a DteRecibidos_db.db
        output_path: ruta del Excel de salida
        only_pending: si True, solo exporta needs_review=1 y SIN_CLASIFICAR

    Returns:
        {"ok": True, "n_rows": int, "path": str}
    """
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row

    # Compatibilidad con DBs antiguas.
    detalle_cols = {r[1] for r in con.execute("PRAGMA table_info(detalle);").fetchall()}
    if "codigo" not in detalle_cols:
        con.execute("ALTER TABLE detalle ADD COLUMN codigo TEXT")
        con.commit()
    if "unidad" not in detalle_cols:
        con.execute("ALTER TABLE detalle ADD COLUMN unidad TEXT")
        con.commit()

    documentos_cols = {r[1] for r in con.execute("PRAGMA table_info(documentos);").fetchall()}
    if "ruta_pdf" not in documentos_cols:
        con.execute("ALTER TABLE documentos ADD COLUMN ruta_pdf TEXT")
        con.commit()

    # --- Catálogo ---
    catalogo = con.execute(
        "SELECT id, categoria_costo, subcategoria_costo, tipo_gasto "
        "FROM catalogo_costos ORDER BY categoria_costo, subcategoria_costo, tipo_gasto"
    ).fetchall()
    categorias_unicas = sorted({r["categoria_costo"] for r in catalogo if r["categoria_costo"]})

    # --- Detalle ---
    where = ""
    if only_pending:
        where = "WHERE d.needs_review = 1 OR d.categoria = 'SIN_CLASIFICAR' OR d.categoria IS NULL"

    detalle_rows = con.execute(f"""
        SELECT
            d.id_det, d.id_doc, d.linea, COALESCE(d.codigo, '') AS codigo, d.descripcion,
            COALESCE(d.unidad, '') AS unidad, d.cantidad,
            d.precio_unitario, d.monto_item,
            COALESCE(d.razon_social, doc.razon_social, '') AS razon_social,
            COALESCE(d.giro, doc.giro, '') AS giro,
            COALESCE(d.fecha_emision, doc.fecha_emision, '') AS fecha_emision,
            COALESCE(doc.ruta_pdf, '') AS ruta_pdf,
            d.categoria, d.subcategoria, d.tipo_gasto,
            d.catalogo_costo_id, d.confianza_categoria, d.needs_review,
            d.origen_clasificacion, d.motivo_clasificacion
        FROM detalle d
        LEFT JOIN documentos doc ON doc.id_doc = d.id_doc
        {where}
        ORDER BY d.fecha_emision DESC, d.id_doc, d.linea
    """).fetchall()

    # --- Documentos ---
    doc_rows = con.execute("""
        SELECT id_doc, tipo_doc, folio, fecha_emision, rut_emisor,
               razon_social, giro, monto_total
        FROM documentos ORDER BY fecha_emision DESC
    """).fetchall()
    con.close()

    # === Crear Excel ===
    wb = Workbook()

    # ── Hoja DETALLE ──
    ws_det = wb.active
    ws_det.title = "detalle"
    headers = _DETALLE_COLS + [_HASH_COL]

    # Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws_det.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER

    # Data
    for row_idx, r in enumerate(detalle_rows, 2):
        cat  = (r["categoria"]    or "").strip()
        sub  = (r["subcategoria"] or "").strip()
        tipo = (r["tipo_gasto"]   or "").strip()
        h    = _row_hash(cat, sub, tipo)

        for col_idx, col_name in enumerate(headers, 1):
            if col_name == _HASH_COL:
                value = h
            else:
                value = r[col_name] if col_name in r.keys() else ""
            cell = ws_det.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _THIN_BORDER

            # Colorear según estado
            nr = r["needs_review"]
            if cat.upper() == "SIN_CLASIFICAR" or not cat:
                cell.fill = _SIN_CLAS_FILL
            elif nr and int(nr) == 1:
                cell.fill = _REVIEW_FILL

            # Columnas no editables en gris
            if col_name not in _EDITABLE_COLS:
                cell.font = _LOCKED_FONT

            # Click en descripcion -> abre PDF del DTE correspondiente
            if col_name == "descripcion":
                pdf_uri = _to_excel_file_uri(r["ruta_pdf"] if "ruta_pdf" in r.keys() else "")
                if pdf_uri:
                    cell.hyperlink = pdf_uri
                    cell.font = _HYPERLINK_FONT

    # Dropdown de categoría
    cat_col_idx = headers.index("categoria") + 1
    cat_formula = '"' + ",".join(categorias_unicas) + '"'
    dv_cat = DataValidation(type="list", formula1=cat_formula, allow_blank=True)
    dv_cat.error = "Selecciona una categoría válida del catálogo"
    dv_cat.errorTitle = "Categoría inválida"
    dv_cat.prompt = "Selecciona la categoría"
    dv_cat.promptTitle = "Categoría"
    ws_det.add_data_validation(dv_cat)
    if len(detalle_rows) > 0:
        col_letter = get_column_letter(cat_col_idx)
        dv_cat.add(f"{col_letter}2:{col_letter}{len(detalle_rows) + 1}")

    # Ocultar columna hash
    hash_col_letter = get_column_letter(len(headers))
    ws_det.column_dimensions[hash_col_letter].hidden = True

    # Auto-ajustar anchos
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row_idx in range(2, min(len(detalle_rows) + 2, 50)):
            cell_val = ws_det.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 40))
        ws_det.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # Filtros
    ws_det.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(detalle_rows) + 1}"
    # Congelar primera fila
    ws_det.freeze_panes = "A2"

    # ── Hoja CATALOGO ──
    ws_cat = wb.create_sheet("catalogo")
    cat_headers = ["id", "categoria_costo", "subcategoria_costo", "tipo_gasto"]
    for col_idx, header in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
    for row_idx, r in enumerate(catalogo, 2):
        for col_idx, col_name in enumerate(cat_headers, 1):
            cell = ws_cat.cell(row=row_idx, column=col_idx, value=r[col_name])
            cell.border = _THIN_BORDER
    ws_cat.freeze_panes = "A2"

    # ── Hoja DOCUMENTOS ──
    ws_doc = wb.create_sheet("documentos")
    doc_headers = ["id_doc", "tipo_doc", "folio", "fecha_emision",
                   "rut_emisor", "razon_social", "giro", "monto_total"]
    for col_idx, header in enumerate(doc_headers, 1):
        cell = ws_doc.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _THIN_BORDER
    for row_idx, r in enumerate(doc_rows, 2):
        for col_idx, col_name in enumerate(doc_headers, 1):
            val = r[col_name] if col_name in r.keys() else ""
            cell = ws_doc.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _THIN_BORDER
    ws_doc.freeze_panes = "A2"

    # Hojas de inventario (stock + entradas)
    n_stock, n_entradas = _append_inventory_sheets(wb, db_path)

    # Guardar
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    return {
        "ok": True,
        "n_rows": len(detalle_rows),
        "n_catalogo": len(catalogo),
        "n_documentos": len(doc_rows),
        "n_stock": n_stock,
        "n_entradas": n_entradas,
        "path": output_path,
    }


# =============================================================================
# IMPORTAR
# =============================================================================

def import_from_excel(
    excel_path: str,
    db_path: str,
    allow_non_category_updates: bool = False,
) -> Dict[str, Any]:
    """
    Lee el Excel, detecta cambios en categoria/subcategoria/tipo_gasto
    y opcionalmente aplica cambios fuera de categoria.

    Returns:
        {"ok": True, "n_changed": int, "changes": [...]}
    """
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    try:
        if "detalle" not in wb.sheetnames:
            return {"ok": False, "error": "No se encontró la hoja 'detalle'"}

        ws = wb["detalle"]

        # Leer headers
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if "id_det" not in headers:
            return {"ok": False, "error": "Columna 'id_det' no encontrada"}

        col_map = {h: i for i, h in enumerate(headers) if h}

        # Validar columnas necesarias
        for needed in ("id_det", "categoria", "subcategoria", "tipo_gasto", _HASH_COL):
            if needed not in col_map:
                return {"ok": False, "error": f"Columna '{needed}' no encontrada"}

        # Leer catálogo válido de la DB
        catalogo_lookup: Dict[str, int] = {}
        for r in con.execute("SELECT id, categoria_costo, subcategoria_costo, tipo_gasto FROM catalogo_costos"):
            key = f"{r['categoria_costo']}|{r['subcategoria_costo']}|{r['tipo_gasto']}"
            catalogo_lookup[key] = int(r["id"])

        detalle_cols = {r[1] for r in con.execute("PRAGMA table_info(detalle);").fetchall()}
        tracked_cols = [c for c in _NON_CATEGORY_SYNC_COLS if c in col_map and c in detalle_cols]
        db_by_id: Dict[str, sqlite3.Row] = {}
        if tracked_cols:
            q_cols = ", ".join(["id_det"] + tracked_cols)
            db_rows = con.execute(f"SELECT {q_cols} FROM detalle").fetchall()
            db_by_id = {str(r["id_det"]): r for r in db_rows}

        # Detectar cambios
        changes: List[Dict[str, Any]] = []
        warnings: List[str] = []
        non_category_changes: List[Dict[str, Any]] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[col_map["id_det"]]:
                continue

            id_det = str(row[col_map["id_det"]]).strip()
            new_cat = (str(row[col_map["categoria"]] or "")).strip().upper()
            new_sub = (str(row[col_map["subcategoria"]] or "")).strip().upper()
            new_tipo = (str(row[col_map["tipo_gasto"]] or "")).strip().upper()
            old_hash = str(row[col_map[_HASH_COL]] or "")

            # Cambios fuera de categoría (preview y/o aplicación opcional)
            db_row = db_by_id.get(id_det)
            if db_row is not None and tracked_cols:
                fields: Dict[str, Any] = {}
                details: List[Dict[str, Any]] = []
                for col in tracked_cols:
                    old_raw = db_row[col] if col in db_row.keys() else None
                    new_raw = row[col_map[col]]
                    if _normalize_for_compare(col, old_raw) != _normalize_for_compare(col, new_raw):
                        new_val = _coerce_for_db(col, new_raw)
                        fields[col] = new_val
                        details.append({"columna": col, "anterior": old_raw, "nuevo": new_val})
                if fields:
                    non_category_changes.append(
                        {
                            "id_det": id_det,
                            "fields": fields,
                            "changes": details,
                        }
                    )

            # Cambios de clasificación
            new_hash = _row_hash(new_cat, new_sub, new_tipo)
            if new_hash == old_hash:
                continue  # Sin cambios de categoria/subcategoria/tipo

            # Validar combinación
            combo_key = f"{new_cat}|{new_sub}|{new_tipo}"
            catalogo_id = catalogo_lookup.get(combo_key)
            if catalogo_id is None:
                # Intentar con tipo_gasto=OTRO
                fallback_key = f"{new_cat}|{new_sub}|OTRO"
                catalogo_id = catalogo_lookup.get(fallback_key)
                if catalogo_id is None:
                    fallback_key2 = f"{new_cat}|OTRO|OTRO"
                    catalogo_id = catalogo_lookup.get(fallback_key2)
                if catalogo_id is not None:
                    # Corregir al valor válido más cercano
                    matched = fallback_key if f"{new_cat}|{new_sub}|OTRO" in catalogo_lookup else fallback_key2
                    parts = matched.split("|")
                    new_sub = parts[1]
                    new_tipo = parts[2]
                    warnings.append(f"{id_det}: ajustado a {matched}")
                else:
                    warnings.append(f"{id_det}: combinación inválida {combo_key}, ignorando")
                    continue

            changes.append(
                {
                    "id_det": id_det,
                    "categoria": new_cat,
                    "subcategoria": new_sub,
                    "tipo_gasto": new_tipo,
                    "catalogo_costo_id": catalogo_id,
                }
            )

        # Aplicar cambios de categoria/subcategoria/tipo
        if changes:
            for ch in changes:
                con.execute(
                    """
                    UPDATE detalle
                    SET categoria = ?,
                        subcategoria = ?,
                        tipo_gasto = ?,
                        catalogo_costo_id = ?,
                        needs_review = 0,
                        origen_clasificacion = 'MANUAL',
                        motivo_clasificacion = ?,
                        confianza_categoria = 100,
                        confianza_subcategoria = 100
                    WHERE id_det = ?
                    """,
                    (
                        ch["categoria"],
                        ch["subcategoria"],
                        ch["tipo_gasto"],
                        ch["catalogo_costo_id"],
                        f"correccion_manual:{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                        ch["id_det"],
                    ),
                )
            con.commit()

        # Aplicar cambios fuera de categoria solo si se confirmó explícitamente.
        n_non_category_applied = 0
        if allow_non_category_updates and non_category_changes:
            for ch in non_category_changes:
                set_parts = []
                vals: List[Any] = []
                for col, val in ch["fields"].items():
                    set_parts.append(f"{col} = ?")
                    vals.append(val)
                vals.append(ch["id_det"])
                con.execute(
                    f"UPDATE detalle SET {', '.join(set_parts)} WHERE id_det = ?",
                    vals,
                )
            con.commit()
            n_non_category_applied = len(non_category_changes)

        return {
            "ok": True,
            "n_changed": len(changes),
            "changes": changes,
            "warnings": warnings,
            "allow_non_category_updates": allow_non_category_updates,
            "n_non_category_changed": len(non_category_changes),
            "n_non_category_applied": n_non_category_applied,
            "non_category_changes": non_category_changes,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass
        con.close()


# =============================================================================
# SYNC COMPLETO: importar + reentrenar
# =============================================================================

def sync_and_retrain(
    excel_path: str,
    db_path: str,
    model_path: Optional[str] = None,
    allow_non_category_updates: bool = False,
) -> Dict[str, Any]:
    """
    Importa cambios del Excel y reentrena el modelo ML si hubo cambios.
    """
    # 1. Importar cambios
    import_result = import_from_excel(
        excel_path,
        db_path,
        allow_non_category_updates=allow_non_category_updates,
    )
    if not import_result.get("ok"):
        return import_result

    n_changed = import_result["n_changed"]
    n_non_category_applied = int(import_result.get("n_non_category_applied", 0) or 0)
    print(
        f"[SYNC] cambios categoria={n_changed} | "
        f"cambios fuera_categoria detectados={import_result.get('n_non_category_changed', 0)} | "
        f"aplicados={n_non_category_applied}"
    )

    if n_changed == 0 and n_non_category_applied == 0:
        return {
            "ok": True,
            "n_changed": 0,
            "n_non_category_changed": import_result.get("n_non_category_changed", 0),
            "n_non_category_applied": 0,
            "retrained": False,
            "message": "Sin cambios detectados",
        }

    # Si solo hubo cambios fuera de categoria, no hace falta reentrenar.
    if n_changed == 0 and n_non_category_applied > 0:
        return {
            "ok": True,
            "n_changed": 0,
            "changes": [],
            "warnings": import_result.get("warnings", []),
            "n_non_category_changed": import_result.get("n_non_category_changed", 0),
            "n_non_category_applied": n_non_category_applied,
            "non_category_changes": import_result.get("non_category_changes", []),
            "retrained": False,
            "message": "Se aplicaron cambios fuera de categoria; no requiere reentrenar.",
        }

    # 2. Reentrenar modelo
    try:
        from app.core.DTE_Recibidos.local_classifier import get_classifier
    except ImportError:
        from local_classifier import get_classifier

    clf = get_classifier(db_path, model_path)
    train_result = clf.retrain()

    return {
        "ok": True,
        "n_changed": n_changed,
        "changes": import_result["changes"],
        "warnings": import_result.get("warnings", []),
        "n_non_category_changed": import_result.get("n_non_category_changed", 0),
        "n_non_category_applied": n_non_category_applied,
        "non_category_changes": import_result.get("non_category_changes", []),
        "retrained": train_result.get("ok", False),
        "train_result": train_result,
    }


# =============================================================================
# RELLENAR columnas desnormalizadas para filas existentes
# =============================================================================

def backfill_denormalized_columns(db_path: str) -> int:
    """
    Rellena razon_social, giro, fecha_emision en detalle
    desde la tabla documentos, para filas que aún no las tienen.
    Retorna cantidad de filas actualizadas.
    """
    con = sqlite3.connect(db_path)
    cur = con.execute("""
        UPDATE detalle
        SET razon_social = (
                SELECT COALESCE(doc.razon_social, '')
                FROM documentos doc WHERE doc.id_doc = detalle.id_doc
            ),
            giro = (
                SELECT COALESCE(doc.giro, '')
                FROM documentos doc WHERE doc.id_doc = detalle.id_doc
            ),
            fecha_emision = (
                SELECT COALESCE(doc.fecha_emision, '')
                FROM documentos doc WHERE doc.id_doc = detalle.id_doc
            )
        WHERE razon_social IS NULL OR razon_social = ''
    """)
    n = cur.rowcount
    con.commit()
    con.close()
    return n


# =============================================================================
# CLI
# =============================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Sincronización Excel <-> SQLite")
    parser.add_argument("action", choices=["export", "import", "sync", "backfill"],
                        help="export: DB→Excel | import: Excel→DB | sync: import+retrain | backfill: llenar columnas")
    parser.add_argument("--db", default=None, help="Ruta a DteRecibidos_db.db")
    parser.add_argument("--excel", default=None, help="Ruta del Excel")
    parser.add_argument("--pending", action="store_true",
                        help="Solo exportar filas pendientes de revisión")
    args = parser.parse_args()

    # Resolver DB desde config.env
    if not args.db:
        base = Path(__file__).resolve().parents[3]
        env_path = base / "data" / "config.env"
        if env_path.exists():
            for line in env_path.read_text(encoding="utf-8").splitlines():
                if line.startswith("DB_PATH_DTE_RECIBIDOS"):
                    args.db = line.split("=", 1)[1].strip().strip("'\"")
                    break

    if not args.db:
        print("[ERROR] Especifica --db o configura DB_PATH_DTE_RECIBIDOS")
        exit(1)

    default_excel = str(Path(args.db).parent / "DteRecibidos_revision.xlsx")
    excel = args.excel or default_excel

    if args.action == "export":
        result = export_to_excel(args.db, excel, only_pending=args.pending)
        print(f"[EXPORT] {result}")

    elif args.action == "import":
        result = import_from_excel(excel, args.db)
        print(f"[IMPORT] {result}")

    elif args.action == "sync":
        result = sync_and_retrain(excel, args.db)
        print(f"[SYNC] {result}")

    elif args.action == "backfill":
        n = backfill_denormalized_columns(args.db)
        print(f"[BACKFILL] {n} filas actualizadas")
