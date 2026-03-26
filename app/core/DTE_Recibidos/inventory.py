"""
Modulo de inventario teorico para Terralix.

Objetivo:
  - Mantener un catalogo maestro por codigo de producto.
  - Registrar movimientos de inventario (entradas desde DTE y salidas manuales por uso).
  - Calcular stock teorico actual por codigo.
"""

from __future__ import annotations

import os
import re
import sqlite3
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

_IGNORED_CODES = {"-"}
_NON_FIELD_CATEGORIES = {"EPP", "MATERIAL", "HERRAMIENTA", "CONTABLE", "RIEGO"}
_NON_FIELD_TYPES = {
    "PROTECCION PERSONAL",
    "INSUMO GENERAL",
    "EQUIPO",
    "NO INVENTARIABLE",
    "INFRAESTRUCTURA",
}
_APP_STATUSES = {"PROGRAMADA", "EJECUTADA", "CANCELADA"}


def _clean_path(path_value: str) -> str:
    return (path_value or "").strip().strip('"').strip("'")


def resolve_db_path(db_path: Optional[str] = None) -> str:
    resolved = _clean_path(db_path or os.getenv("DB_PATH_DTE_RECIBIDOS", ""))
    if not resolved:
        raise ValueError("No se encontro DB_PATH_DTE_RECIBIDOS en variables de entorno.")
    return resolved


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_code(value: Any) -> str:
    return _normalize_text(value).upper()


def _is_valid_inventory_code(value: Any) -> bool:
    code = _normalize_code(value)
    return bool(code) and code not in _IGNORED_CODES


def _is_field_product(categoria: Any, tipo: Any) -> bool:
    cat = _normalize_text(categoria).upper()
    tpo = _normalize_text(tipo).upper()
    if cat in _NON_FIELD_CATEGORIES:
        return False
    if tpo in _NON_FIELD_TYPES:
        return False
    return True


def _normalize_header(value: Any) -> str:
    s = _normalize_text(value)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s).strip().upper()
    return s


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    s = _normalize_text(value)
    if not s:
        return 0.0

    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif re.fullmatch(r"\d{1,3}(?:\.\d{3})+", s):
        s = s.replace(".", "")

    try:
        return float(s)
    except Exception:
        return 0.0


def _parse_float_strict(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    s = _normalize_text(value)
    if not s:
        return None

    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif re.fullmatch(r"\d{1,3}(?:\.\d{3})+", s):
        s = s.replace(".", "")

    try:
        return float(s)
    except Exception:
        return None


def _to_int(value: Any) -> int:
    try:
        return int(round(_to_float(value)))
    except Exception:
        return 0


def _coerce_date_iso(value: Any) -> str:
    if value is None:
        return date.today().isoformat()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    raw = _normalize_text(value)
    if not raw:
        return date.today().isoformat()

    # Casos comunes: YYYY-MM-DD, YYYY-MM-DD HH:MM:SS
    raw10 = raw[:10]
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw10, fmt).date().isoformat()
        except Exception:
            pass

    # Fallback: hoy
    return date.today().isoformat()


def _parse_date_iso_strict(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    raw = _normalize_text(value)
    if not raw:
        return None

    raw10 = raw[:10]
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw10, fmt).date().isoformat()
        except Exception:
            pass
    return None


def _normalize_application_status(value: Any) -> str:
    raw = _normalize_text(value).upper()
    aliases = {
        "": "PROGRAMADA",
        "P": "PROGRAMADA",
        "PROGRAMADA": "PROGRAMADA",
        "PROG": "PROGRAMADA",
        "E": "EJECUTADA",
        "EJECUTADA": "EJECUTADA",
        "REALIZADA": "EJECUTADA",
        "COMPLETADA": "EJECUTADA",
        "C": "CANCELADA",
        "CANCELADA": "CANCELADA",
        "ANULADA": "CANCELADA",
    }
    return aliases.get(raw, raw)


def _connect(db_path: str) -> sqlite3.Connection:
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    return con


def _add_column_if_missing(con: sqlite3.Connection, table: str, column: str, coltype: str) -> None:
    cols = {r[1] for r in con.execute(f"PRAGMA table_info({table});").fetchall()}
    if column not in cols:
        con.execute(f"ALTER TABLE {table} ADD COLUMN {column} {coltype}")


def _purge_ignored_codes_in_connection(con: sqlite3.Connection) -> Dict[str, int]:
    n_mov = con.execute(
        "DELETE FROM inventario_movimientos WHERE UPPER(TRIM(COALESCE(codigo, ''))) IN ('', '-')"
    ).rowcount
    n_cat = con.execute(
        "DELETE FROM inventario_catalogo WHERE UPPER(TRIM(COALESCE(codigo, ''))) IN ('', '-')"
    ).rowcount
    return {"catalog_deleted": int(n_cat or 0), "movements_deleted": int(n_mov or 0)}


def ensure_inventory_schema(db_path: str) -> None:
    con = _connect(db_path)
    try:
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS inventario_catalogo (
                codigo TEXT PRIMARY KEY,
                descripcion_estandar TEXT,
                unidad_base TEXT,
                categoria TEXT,
                tipo TEXT,
                ocurrencias INTEGER DEFAULT 0,
                variaciones TEXT DEFAULT 'NO',
                activo INTEGER DEFAULT 1,
                fuente TEXT,
                updated_at TEXT DEFAULT (datetime('now'))
            )
            """
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS inventario_movimientos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT NOT NULL,
                fecha TEXT NOT NULL,
                tipo_mov TEXT NOT NULL,
                cantidad REAL NOT NULL,
                unidad TEXT,
                signo INTEGER NOT NULL,
                fuente TEXT NOT NULL,
                referencia TEXT,
                observacion TEXT,
                source_hash TEXT UNIQUE,
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (codigo) REFERENCES inventario_catalogo(codigo)
            )
            """
        )
        con.execute(
            "CREATE INDEX IF NOT EXISTS idx_inv_mov_codigo_fecha ON inventario_movimientos(codigo, fecha)"
        )
        con.execute(
            "CREATE INDEX IF NOT EXISTS idx_inv_mov_fuente ON inventario_movimientos(fuente)"
        )
        con.execute(
            "CREATE INDEX IF NOT EXISTS idx_inv_catalogo_categoria ON inventario_catalogo(categoria, tipo)"
        )
        _add_column_if_missing(con, "inventario_catalogo", "ultima_fecha_override", "TEXT")
        _add_column_if_missing(con, "inventario_catalogo", "ultima_tipo_override", "TEXT")
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS aplicaciones_campo (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                titulo TEXT NOT NULL,
                descripcion TEXT,
                fecha_programada TEXT NOT NULL,
                fecha_ejecucion TEXT,
                estado TEXT NOT NULL DEFAULT 'PROGRAMADA',
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now'))
            )
            """
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS aplicaciones_campo_productos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                aplicacion_id INTEGER NOT NULL,
                codigo TEXT NOT NULL,
                cantidad REAL NOT NULL,
                unidad TEXT,
                observacion TEXT,
                movimiento_id INTEGER,
                created_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (aplicacion_id) REFERENCES aplicaciones_campo(id),
                FOREIGN KEY (codigo) REFERENCES inventario_catalogo(codigo),
                FOREIGN KEY (movimiento_id) REFERENCES inventario_movimientos(id)
            )
            """
        )
        con.execute(
            "CREATE INDEX IF NOT EXISTS idx_app_campo_fecha_estado ON aplicaciones_campo(fecha_programada, estado)"
        )
        con.execute(
            "CREATE INDEX IF NOT EXISTS idx_app_prod_aplicacion ON aplicaciones_campo_productos(aplicacion_id)"
        )
        con.execute(
            "CREATE INDEX IF NOT EXISTS idx_app_prod_movimiento ON aplicaciones_campo_productos(movimiento_id)"
        )
        _add_column_if_missing(con, "aplicaciones_campo", "titulo", "TEXT")
        _add_column_if_missing(con, "aplicaciones_campo", "descripcion", "TEXT")
        _add_column_if_missing(con, "aplicaciones_campo", "fecha_programada", "TEXT")
        _add_column_if_missing(con, "aplicaciones_campo", "fecha_ejecucion", "TEXT")
        _add_column_if_missing(con, "aplicaciones_campo", "estado", "TEXT")
        _add_column_if_missing(con, "aplicaciones_campo", "updated_at", "TEXT")
        _add_column_if_missing(con, "aplicaciones_campo_productos", "unidad", "TEXT")
        _add_column_if_missing(con, "aplicaciones_campo_productos", "observacion", "TEXT")
        _add_column_if_missing(con, "aplicaciones_campo_productos", "movimiento_id", "INTEGER")
        _purge_ignored_codes_in_connection(con)
        con.commit()
    finally:
        con.close()


def _header_aliases() -> Dict[str, set[str]]:
    return {
        "codigo": {"CODIGO", "COD", "COD.", "SKU", "ITEM"},
        "descripcion_estandar": {
            "DESCRIPCION ESTANDAR",
            "DESCRIPCION",
            "PRODUCTO",
            "INSUMO",
        },
        "unidad_base": {"UNIDAD", "UND", "UM"},
        "categoria": {"CATEGORIA"},
        "tipo": {"TIPO"},
        "ocurrencias": {"OCURRENCIAS"},
        "variaciones": {"VARIACIONES"},
    }


def import_catalog_from_excel(
    excel_path: str,
    db_path: str,
    source_name: str = "EXCEL_MAESTRO",
) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return {"ok": False, "error": "Falta la dependencia openpyxl para leer Excel."}

    ensure_inventory_schema(db_path)

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    try:
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        wb.close()
        return {"ok": False, "error": "El Excel de inventario esta vacio."}

    alias = _header_aliases()
    col_map: Dict[str, int] = {}
    for idx, h in enumerate(headers):
        norm = _normalize_header(h)
        for target, options in alias.items():
            if norm in options and target not in col_map:
                col_map[target] = idx

    if "codigo" not in col_map:
        wb.close()
        return {"ok": False, "error": "No se encontro la columna 'Codigo' en el Excel."}

    con = _connect(db_path)
    inserted = 0
    updated = 0
    skipped = 0

    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue

            codigo = _normalize_code(row[col_map["codigo"]])
            if not _is_valid_inventory_code(codigo):
                skipped += 1
                continue

            desc = _normalize_text(row[col_map["descripcion_estandar"]]) if "descripcion_estandar" in col_map else ""
            unidad = _normalize_text(row[col_map["unidad_base"]]).upper() if "unidad_base" in col_map else ""
            categoria = _normalize_text(row[col_map["categoria"]]).upper() if "categoria" in col_map else ""
            tipo = _normalize_text(row[col_map["tipo"]]).upper() if "tipo" in col_map else ""
            ocurrencias = _to_int(row[col_map["ocurrencias"]]) if "ocurrencias" in col_map else 0
            variaciones = _normalize_text(row[col_map["variaciones"]]).upper() if "variaciones" in col_map else "NO"

            old = con.execute(
                """
                SELECT codigo, descripcion_estandar, unidad_base, categoria, tipo, ocurrencias, variaciones
                FROM inventario_catalogo
                WHERE codigo = ?
                """,
                (codigo,),
            ).fetchone()

            if old is None:
                con.execute(
                    """
                    INSERT INTO inventario_catalogo
                    (codigo, descripcion_estandar, unidad_base, categoria, tipo, ocurrencias, variaciones, activo, fuente, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?, datetime('now'))
                    """,
                    (
                        codigo,
                        desc,
                        unidad,
                        categoria,
                        tipo,
                        ocurrencias,
                        variaciones or "NO",
                        source_name,
                    ),
                )
                inserted += 1
                continue

            new_desc = desc or (old["descripcion_estandar"] or "")
            new_unidad = unidad or (old["unidad_base"] or "")
            new_categoria = categoria or (old["categoria"] or "")
            new_tipo = tipo or (old["tipo"] or "")
            new_ocurr = ocurrencias if ocurrencias > 0 else int(old["ocurrencias"] or 0)
            new_var = variaciones or (old["variaciones"] or "NO")

            changed = (
                new_desc != (old["descripcion_estandar"] or "")
                or new_unidad != (old["unidad_base"] or "")
                or new_categoria != (old["categoria"] or "")
                or new_tipo != (old["tipo"] or "")
                or new_ocurr != int(old["ocurrencias"] or 0)
                or new_var != (old["variaciones"] or "NO")
            )

            if changed:
                con.execute(
                    """
                    UPDATE inventario_catalogo
                    SET descripcion_estandar = ?,
                        unidad_base = ?,
                        categoria = ?,
                        tipo = ?,
                        ocurrencias = ?,
                        variaciones = ?,
                        activo = 1,
                        fuente = ?,
                        updated_at = datetime('now')
                    WHERE codigo = ?
                    """,
                    (
                        new_desc,
                        new_unidad,
                        new_categoria,
                        new_tipo,
                        new_ocurr,
                        new_var,
                        source_name,
                        codigo,
                    ),
                )
                updated += 1
            else:
                skipped += 1

        con.commit()
    finally:
        wb.close()
        con.close()

    return {
        "ok": True,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "total_processed": inserted + updated + skipped,
    }


def _table_exists(con: sqlite3.Connection, table_name: str) -> bool:
    row = con.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone()
    return row is not None


def _refresh_catalog_stats_from_detalle(
    con: sqlite3.Connection,
    only_categoria: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Recalcula ocurrencias/variaciones por codigo usando detalle.
    Respeta la exclusion de codigos no inventariables ('-') y usa solo Facturas.
    """
    if not _table_exists(con, "detalle"):
        return {"ok": False, "error": "No existe la tabla detalle"}

    detalle_cols = {r[1] for r in con.execute("PRAGMA table_info(detalle);").fetchall()}
    if "codigo" not in detalle_cols:
        return {"ok": False, "error": "La tabla detalle no tiene columna codigo"}

    select_parts = [
        "UPPER(TRIM(COALESCE(d.codigo, ''))) AS codigo",
        "UPPER(TRIM(COALESCE(d.descripcion, ''))) AS descripcion_norm" if "descripcion" in detalle_cols else "'' AS descripcion_norm",
        "UPPER(TRIM(COALESCE(d.unidad, ''))) AS unidad_norm" if "unidad" in detalle_cols else "'' AS unidad_norm",
        "UPPER(TRIM(COALESCE(d.categoria, ''))) AS categoria_norm" if "categoria" in detalle_cols else "'' AS categoria_norm",
        "COUNT(*) AS ocurrencias_total",
        "COUNT(DISTINCT UPPER(TRIM(COALESCE(d.descripcion, '')))) AS n_desc" if "descripcion" in detalle_cols else "1 AS n_desc",
        "COUNT(DISTINCT UPPER(TRIM(COALESCE(d.unidad, '')))) AS n_uni" if "unidad" in detalle_cols else "1 AS n_uni",
    ]
    where_parts = [
        "TRIM(COALESCE(d.codigo, '')) <> ''",
        "UPPER(TRIM(COALESCE(d.codigo, ''))) <> '-'",
    ]
    if "id_doc" in detalle_cols:
        where_parts.append("COALESCE(d.id_doc, '') LIKE 'Factura_%'")
    params: List[Any] = []
    if only_categoria and "categoria" in detalle_cols:
        where_parts.append("UPPER(TRIM(COALESCE(d.categoria, ''))) = ?")
        params.append(only_categoria.strip().upper())

    # Se toma descripcion/unidad/categoria con mayor frecuencia por codigo.
    grouped = con.execute(
        f"""
        WITH base AS (
            SELECT
                UPPER(TRIM(COALESCE(d.codigo, ''))) AS codigo,
                UPPER(TRIM(COALESCE(d.descripcion, ''))) AS descripcion_norm,
                UPPER(TRIM(COALESCE(d.unidad, ''))) AS unidad_norm,
                UPPER(TRIM(COALESCE(d.categoria, ''))) AS categoria_norm
            FROM detalle d
            WHERE {' AND '.join(where_parts)}
        ),
        desc_mode AS (
            SELECT codigo, descripcion_norm
            FROM (
                SELECT
                    codigo,
                    descripcion_norm,
                    COUNT(*) AS c,
                    ROW_NUMBER() OVER (
                        PARTITION BY codigo
                        ORDER BY COUNT(*) DESC, descripcion_norm
                    ) AS rn
                FROM base
                WHERE descripcion_norm <> ''
                GROUP BY codigo, descripcion_norm
            ) x
            WHERE rn = 1
        ),
        uni_mode AS (
            SELECT codigo, unidad_norm
            FROM (
                SELECT
                    codigo,
                    unidad_norm,
                    COUNT(*) AS c,
                    ROW_NUMBER() OVER (
                        PARTITION BY codigo
                        ORDER BY COUNT(*) DESC, unidad_norm
                    ) AS rn
                FROM base
                WHERE unidad_norm <> ''
                GROUP BY codigo, unidad_norm
            ) x
            WHERE rn = 1
        ),
        cat_mode AS (
            SELECT codigo, categoria_norm
            FROM (
                SELECT
                    codigo,
                    categoria_norm,
                    COUNT(*) AS c,
                    ROW_NUMBER() OVER (
                        PARTITION BY codigo
                        ORDER BY COUNT(*) DESC, categoria_norm
                    ) AS rn
                FROM base
                WHERE categoria_norm <> ''
                GROUP BY codigo, categoria_norm
            ) x
            WHERE rn = 1
        ),
        agg AS (
            SELECT
                codigo,
                COUNT(*) AS ocurrencias_total,
                COUNT(DISTINCT CASE WHEN descripcion_norm <> '' THEN descripcion_norm END) AS n_desc,
                COUNT(DISTINCT CASE WHEN unidad_norm <> '' THEN unidad_norm END) AS n_uni
            FROM base
            GROUP BY codigo
        )
        SELECT
            a.codigo AS codigo,
            COALESCE(dm.descripcion_norm, '') AS descripcion_mode,
            COALESCE(um.unidad_norm, '') AS unidad_mode,
            COALESCE(cm.categoria_norm, '') AS categoria_mode,
            a.ocurrencias_total AS ocurrencias_total,
            a.n_desc AS n_desc,
            a.n_uni AS n_uni
        FROM agg a
        LEFT JOIN desc_mode dm ON dm.codigo = a.codigo
        LEFT JOIN uni_mode um ON um.codigo = a.codigo
        LEFT JOIN cat_mode cm ON cm.codigo = a.codigo
        """,
        params,
    ).fetchall()

    updated = 0
    inserted = 0
    variaciones_si = 0
    for r in grouped:
        codigo = _normalize_code(r["codigo"])
        if not _is_valid_inventory_code(codigo):
            continue

        desc_mode = _normalize_text(r["descripcion_mode"]).upper()
        unidad_mode = _normalize_text(r["unidad_mode"]).upper()
        categoria_mode = _normalize_text(r["categoria_mode"]).upper()
        ocurrencias = _to_int(r["ocurrencias_total"])
        n_desc = _to_int(r["n_desc"])
        n_uni = _to_int(r["n_uni"])
        variaciones = "SI" if n_desc > 1 or n_uni > 1 else "NO"
        if variaciones == "SI":
            variaciones_si += 1

        old = con.execute(
            """
            SELECT codigo, descripcion_estandar, unidad_base, categoria, tipo, ocurrencias, variaciones
            FROM inventario_catalogo
            WHERE codigo = ?
            """,
            (codigo,),
        ).fetchone()

        if old is None:
            con.execute(
                """
                INSERT INTO inventario_catalogo
                (codigo, descripcion_estandar, unidad_base, categoria, tipo, ocurrencias, variaciones, activo, fuente, updated_at)
                VALUES (?, ?, ?, ?, '', ?, ?, 1, 'AUTO_DETALLE', datetime('now'))
                """,
                (codigo, desc_mode, unidad_mode, categoria_mode, ocurrencias, variaciones),
            )
            inserted += 1
            continue

        new_desc = old["descripcion_estandar"] or desc_mode
        new_unidad = old["unidad_base"] or unidad_mode
        new_categoria = old["categoria"] or categoria_mode
        new_tipo = old["tipo"] or ""
        new_ocurr = ocurrencias
        new_var = variaciones

        if (
            new_desc != (old["descripcion_estandar"] or "")
            or new_unidad != (old["unidad_base"] or "")
            or new_categoria != (old["categoria"] or "")
            or new_tipo != (old["tipo"] or "")
            or new_ocurr != int(old["ocurrencias"] or 0)
            or new_var != (old["variaciones"] or "NO")
        ):
            con.execute(
                """
                UPDATE inventario_catalogo
                SET descripcion_estandar = ?,
                    unidad_base = ?,
                    categoria = ?,
                    tipo = ?,
                    ocurrencias = ?,
                    variaciones = ?,
                    activo = 1,
                    fuente = CASE
                        WHEN COALESCE(fuente, '') = '' THEN 'AUTO_DETALLE'
                        ELSE fuente
                    END,
                    updated_at = datetime('now')
                WHERE codigo = ?
                """,
                (
                    new_desc,
                    new_unidad,
                    new_categoria,
                    new_tipo,
                    new_ocurr,
                    new_var,
                    codigo,
                ),
            )
            updated += 1

    return {
        "ok": True,
        "codes_seen": len(grouped),
        "catalog_inserted_stats": inserted,
        "catalog_updated_stats": updated,
        "codes_variaciones_si": variaciones_si,
    }


def sync_entries_from_detalle(
    db_path: str,
    only_categoria: Optional[str] = None,
) -> Dict[str, Any]:
    ensure_inventory_schema(db_path)
    con = _connect(db_path)
    try:
        if not _table_exists(con, "detalle"):
            return {"ok": False, "error": "No existe la tabla 'detalle' en la base de datos."}

        detalle_cols = {r[1] for r in con.execute("PRAGMA table_info(detalle);").fetchall()}
        required = {"id_det", "id_doc", "codigo", "descripcion", "cantidad"}
        missing = sorted(required - detalle_cols)
        if missing:
            return {"ok": False, "error": f"Faltan columnas en detalle: {', '.join(missing)}"}

        removed_non_factura_movements = con.execute(
            """
            DELETE FROM inventario_movimientos
            WHERE fuente = 'DTE_DETALLE'
              AND (
                    COALESCE(source_hash, '') = ''
                 OR COALESCE(source_hash, '') NOT LIKE 'DTE:Factura_%'
              )
            """
        ).rowcount

        select_parts = [
            "d.id_det",
            "COALESCE(d.id_doc, '') AS id_doc",
            "COALESCE(d.codigo, '') AS codigo",
            "COALESCE(d.descripcion, '') AS descripcion",
            "COALESCE(d.cantidad, 0) AS cantidad",
            "COALESCE(d.unidad, '') AS unidad" if "unidad" in detalle_cols else "'' AS unidad",
            "COALESCE(d.categoria, '') AS categoria" if "categoria" in detalle_cols else "'' AS categoria",
            "COALESCE(d.fecha_emision, '') AS fecha_emision" if "fecha_emision" in detalle_cols else "'' AS fecha_emision",
        ]

        query = f"""
            SELECT {", ".join(select_parts)}
            FROM detalle d
            WHERE TRIM(COALESCE(d.codigo, '')) <> ''
              AND UPPER(TRIM(COALESCE(d.codigo, ''))) <> '-'
              AND COALESCE(d.cantidad, 0) > 0
              AND COALESCE(d.id_doc, '') LIKE 'Factura_%'
        """
        params: List[Any] = []
        if only_categoria and "categoria" in detalle_cols:
            query += " AND UPPER(TRIM(COALESCE(d.categoria, ''))) = ?"
            params.append(only_categoria.strip().upper())

        rows = con.execute(query, params).fetchall()

        mov_inserted = 0
        mov_updated = 0
        cat_inserted = 0
        cat_updated = 0

        for r in rows:
            codigo = _normalize_code(r["codigo"])
            if not _is_valid_inventory_code(codigo):
                continue

            descripcion = _normalize_text(r["descripcion"])
            unidad = _normalize_text(r["unidad"]).upper()
            categoria = _normalize_text(r["categoria"]).upper()
            cantidad = _to_float(r["cantidad"])
            fecha = _coerce_date_iso(r["fecha_emision"])
            id_det = _normalize_text(r["id_det"])

            old_cat = con.execute(
                "SELECT codigo, descripcion_estandar, unidad_base, categoria FROM inventario_catalogo WHERE codigo = ?",
                (codigo,),
            ).fetchone()

            if old_cat is None:
                con.execute(
                    """
                    INSERT INTO inventario_catalogo
                    (codigo, descripcion_estandar, unidad_base, categoria, tipo, ocurrencias, variaciones, activo, fuente, updated_at)
                    VALUES (?, ?, ?, ?, '', 0, 'NO', 1, 'SYNC_DTE', datetime('now'))
                    """,
                    (codigo, descripcion, unidad, categoria),
                )
                cat_inserted += 1
            else:
                new_desc = old_cat["descripcion_estandar"] or descripcion
                new_unidad = old_cat["unidad_base"] or unidad
                new_categoria = old_cat["categoria"] or categoria
                if (
                    new_desc != (old_cat["descripcion_estandar"] or "")
                    or new_unidad != (old_cat["unidad_base"] or "")
                    or new_categoria != (old_cat["categoria"] or "")
                ):
                    con.execute(
                        """
                        UPDATE inventario_catalogo
                        SET descripcion_estandar = ?,
                            unidad_base = ?,
                            categoria = ?,
                            updated_at = datetime('now')
                        WHERE codigo = ?
                        """,
                        (new_desc, new_unidad, new_categoria, codigo),
                    )
                    cat_updated += 1

            source_hash = f"DTE:{id_det}"
            old_mov = con.execute(
                "SELECT id FROM inventario_movimientos WHERE source_hash = ?",
                (source_hash,),
            ).fetchone()

            if old_mov is None:
                con.execute(
                    """
                    INSERT INTO inventario_movimientos
                    (codigo, fecha, tipo_mov, cantidad, unidad, signo, fuente, referencia, observacion, source_hash)
                    VALUES (?, ?, 'ENTRADA_COMPRA', ?, ?, 1, 'DTE_DETALLE', ?, ?, ?)
                    """,
                    (
                        codigo,
                        fecha,
                        cantidad,
                        unidad,
                        id_det,
                        descripcion,
                        source_hash,
                    ),
                )
                mov_inserted += 1
            else:
                con.execute(
                    """
                    UPDATE inventario_movimientos
                    SET codigo = ?,
                        fecha = ?,
                        tipo_mov = 'ENTRADA_COMPRA',
                        cantidad = ?,
                        unidad = ?,
                        signo = 1,
                        fuente = 'DTE_DETALLE',
                        referencia = ?,
                        observacion = ?
                    WHERE source_hash = ?
                    """,
                    (
                        codigo,
                        fecha,
                        cantidad,
                        unidad,
                        id_det,
                        descripcion,
                        source_hash,
                    ),
                )
                mov_updated += 1

        stats = _refresh_catalog_stats_from_detalle(
            con=con,
            only_categoria=only_categoria,
        )
        purge = _purge_ignored_codes_in_connection(con)
        con.commit()
        return {
            "ok": True,
            "rows_scanned": len(rows),
            "movements_inserted": mov_inserted,
            "movements_updated": mov_updated,
            "catalog_inserted": cat_inserted,
            "catalog_updated": cat_updated,
            "catalog_stats": stats,
            "purge_ignored_codes": purge,
            "removed_non_factura_movements": int(removed_non_factura_movements or 0),
            "categoria_filtro": only_categoria or "(todas)",
            "doc_tipo_filtro": "Factura_*",
        }
    finally:
        con.close()


def list_catalog_products(db_path: str) -> List[Dict[str, Any]]:
    ensure_inventory_schema(db_path)
    con = _connect(db_path)
    try:
        rows = con.execute(
            """
            SELECT codigo, descripcion_estandar, unidad_base, categoria, tipo, activo
            FROM inventario_catalogo
            WHERE COALESCE(activo, 1) = 1
              AND UPPER(TRIM(COALESCE(codigo, ''))) NOT IN ('', '-')
              AND UPPER(TRIM(COALESCE(categoria, ''))) NOT IN ('EPP', 'MATERIAL', 'HERRAMIENTA', 'CONTABLE', 'RIEGO')
              AND UPPER(TRIM(COALESCE(tipo, ''))) NOT IN (
                    'PROTECCION PERSONAL',
                    'INSUMO GENERAL',
                    'EQUIPO',
                    'NO INVENTARIABLE',
                    'INFRAESTRUCTURA'
              )
            ORDER BY categoria, tipo, codigo
            """
        ).fetchall()
        return [dict(r) for r in rows if _is_field_product(r["categoria"], r["tipo"])]
    finally:
        con.close()


def _stock_for_code(con: sqlite3.Connection, codigo: str) -> float:
    row = con.execute(
        """
        SELECT COALESCE(SUM(signo * cantidad), 0) AS stock
        FROM inventario_movimientos
        WHERE codigo = ?
        """,
        (codigo,),
    ).fetchone()
    return float(row["stock"] if row else 0.0)


def _register_usage_in_connection(
    con: sqlite3.Connection,
    codigo: str,
    cantidad: float,
    fecha_uso: Any,
    observacion: str = "",
    unidad: str = "",
    allow_negative: bool = False,
    fuente: str = "MANUAL_USO",
    referencia: Optional[str] = None,
) -> Dict[str, Any]:
    codigo_norm = _normalize_code(codigo)
    if not _is_valid_inventory_code(codigo_norm):
        return {"ok": False, "error": "Codigo invalido para inventario (vacio o '-')."}

    qty = _to_float(cantidad)
    if qty <= 0:
        return {"ok": False, "error": "La cantidad debe ser mayor a 0."}

    fecha = _coerce_date_iso(fecha_uso)
    obs = _normalize_text(observacion)
    um = _normalize_text(unidad).upper()

    cat = con.execute(
        "SELECT codigo, unidad_base FROM inventario_catalogo WHERE codigo = ?",
        (codigo_norm,),
    ).fetchone()
    if cat is None:
        con.execute(
            """
            INSERT INTO inventario_catalogo
            (codigo, descripcion_estandar, unidad_base, categoria, tipo, ocurrencias, variaciones, activo, fuente, updated_at)
            VALUES (?, '', ?, '', '', 0, 'NO', 1, ?, datetime('now'))
            """,
            (codigo_norm, um, fuente),
        )
        unidad_final = um
    else:
        unidad_final = um or _normalize_text(cat["unidad_base"]).upper()

    stock_before = _stock_for_code(con, codigo_norm)
    stock_after = stock_before - qty
    if stock_after < 0 and not allow_negative:
        return {
            "ok": False,
            "error": "El movimiento deja stock negativo.",
            "stock_before": stock_before,
            "stock_after": stock_after,
            "codigo": codigo_norm,
        }

    cur = con.execute(
        """
        INSERT INTO inventario_movimientos
        (codigo, fecha, tipo_mov, cantidad, unidad, signo, fuente, referencia, observacion, source_hash)
        VALUES (?, ?, 'SALIDA_USO', ?, ?, -1, ?, ?, ?, NULL)
        """,
        (
            codigo_norm,
            fecha,
            qty,
            unidad_final,
            _normalize_text(fuente) or "MANUAL_USO",
            _normalize_text(referencia) or None,
            obs,
        ),
    )

    return {
        "ok": True,
        "movement_id": cur.lastrowid,
        "codigo": codigo_norm,
        "cantidad": qty,
        "stock_before": stock_before,
        "stock_after": stock_after,
        "unidad": unidad_final,
        "fecha": fecha,
    }


def register_usage(
    db_path: str,
    codigo: str,
    cantidad: float,
    fecha_uso: Any,
    observacion: str = "",
    unidad: str = "",
    allow_negative: bool = False,
) -> Dict[str, Any]:
    ensure_inventory_schema(db_path)
    con = _connect(db_path)
    try:
        result = _register_usage_in_connection(
            con=con,
            codigo=codigo,
            cantidad=cantidad,
            fecha_uso=fecha_uso,
            observacion=observacion,
            unidad=unidad,
            allow_negative=allow_negative,
            fuente="MANUAL_USO",
            referencia=None,
        )
        if not result.get("ok"):
            return result
        con.commit()
        return result
    finally:
        con.close()


def update_stock_cell(
    db_path: str,
    codigo: str,
    column_name: str,
    new_value: Any,
    observacion: str = "EDICION_CELDA",
) -> Dict[str, Any]:
    """
    Edita una celda del modulo de stock (excepto codigo).

    Columnas soportadas:
      - descripcion_estandar: actualiza catalogo
      - unidad_base: actualiza catalogo
      - stock_actual: crea movimiento de ajuste por delta
      - ultima_fecha: guarda override visual sin tocar movimientos historicos
      - ultima_modificacion_tipo: guarda override visual sin tocar movimientos historicos
    """
    ensure_inventory_schema(db_path)

    codigo_norm = _normalize_code(codigo)
    if not _is_valid_inventory_code(codigo_norm):
        return {"ok": False, "error": "Codigo invalido para inventario."}

    col = _normalize_text(column_name)
    if col == "codigo":
        return {"ok": False, "error": "La columna 'codigo' no es editable."}

    editable_cols = {
        "descripcion_estandar",
        "unidad_base",
        "stock_actual",
        "ultima_fecha",
        "ultima_modificacion_tipo",
    }
    if col not in editable_cols:
        return {"ok": False, "error": f"Columna no editable: {col}"}

    con = _connect(db_path)
    try:
        # 1) Columnas maestras del catalogo
        if col in {"descripcion_estandar", "unidad_base"}:
            val = _normalize_text(new_value)
            if col == "unidad_base":
                val = val.upper()

            old_cat = con.execute(
                """
                SELECT codigo, descripcion_estandar, unidad_base
                FROM inventario_catalogo
                WHERE codigo = ?
                """,
                (codigo_norm,),
            ).fetchone()

            if old_cat is None:
                desc = val if col == "descripcion_estandar" else ""
                unidad = val if col == "unidad_base" else ""
                con.execute(
                    """
                    INSERT INTO inventario_catalogo
                    (codigo, descripcion_estandar, unidad_base, categoria, tipo, ocurrencias, variaciones, activo, fuente, updated_at)
                    VALUES (?, ?, ?, '', '', 0, 'NO', 1, 'MANUAL_EDICION_CELDA', datetime('now'))
                    """,
                    (codigo_norm, desc, unidad),
                )
                before = ""
            else:
                before = _normalize_text(old_cat[col])
                con.execute(
                    f"""
                    UPDATE inventario_catalogo
                    SET {col} = ?,
                        updated_at = datetime('now')
                    WHERE codigo = ?
                    """,
                    (val, codigo_norm),
                )

            con.commit()
            return {
                "ok": True,
                "action": "catalog_update",
                "codigo": codigo_norm,
                "column": col,
                "before": before,
                "after": val,
            }

        # 2) Stock actual -> ajuste por delta
        if col == "stock_actual":
            parsed = _parse_float_strict(new_value)
            if parsed is None:
                return {"ok": False, "error": "Stock invalido. Ingresa un numero valido."}
            target = float(parsed)
            stock_before = _stock_for_code(con, codigo_norm)
            delta = round(target - stock_before, 6)

            if abs(delta) < 1e-9:
                return {
                    "ok": True,
                    "action": "stock_adjust",
                    "codigo": codigo_norm,
                    "stock_before": stock_before,
                    "stock_after": stock_before,
                    "delta": 0.0,
                    "no_change": True,
                }

            cat = con.execute(
                "SELECT unidad_base FROM inventario_catalogo WHERE codigo = ?",
                (codigo_norm,),
            ).fetchone()
            unidad = _normalize_text(cat["unidad_base"]) if cat else ""
            sign = 1 if delta > 0 else -1
            qty = abs(delta)
            tipo_mov = "AJUSTE_ENTRADA_MANUAL" if sign > 0 else "AJUSTE_SALIDA_MANUAL"

            cur = con.execute(
                """
                INSERT INTO inventario_movimientos
                (codigo, fecha, tipo_mov, cantidad, unidad, signo, fuente, referencia, observacion, source_hash)
                VALUES (?, ?, ?, ?, ?, ?, 'MANUAL_EDICION_CELDA', NULL, ?, NULL)
                """,
                (
                    codigo_norm,
                    date.today().isoformat(),
                    tipo_mov,
                    qty,
                    unidad,
                    sign,
                    f"{observacion}:stock_actual",
                ),
            )
            con.commit()
            stock_after = _stock_for_code(con, codigo_norm)
            return {
                "ok": True,
                "action": "stock_adjust",
                "movement_id": cur.lastrowid,
                "codigo": codigo_norm,
                "stock_before": stock_before,
                "stock_after": stock_after,
                "delta": delta,
            }

        # 3) Fecha de la ultima modificacion (override visual, no historico)
        if col == "ultima_fecha":
            raw = _normalize_text(new_value)
            new_override = _parse_date_iso_strict(raw) if raw else None
            if raw and not new_override:
                return {
                    "ok": False,
                    "error": "Fecha invalida. Usa formato YYYY-MM-DD (o DD/MM/YYYY).",
                }

            cat = con.execute(
                """
                SELECT codigo, ultima_fecha_override
                FROM inventario_catalogo
                WHERE codigo = ?
                """,
                (codigo_norm,),
            ).fetchone()
            last_mov = con.execute(
                """
                SELECT fecha
                FROM inventario_movimientos
                WHERE codigo = ?
                ORDER BY fecha DESC, id DESC
                LIMIT 1
                """,
                (codigo_norm,),
            ).fetchone()

            old_override = _parse_date_iso_strict(cat["ultima_fecha_override"]) if cat else None
            last_date = _normalize_text(last_mov["fecha"]) if last_mov else ""
            before = old_override or last_date
            after = new_override or last_date

            if old_override == new_override:
                return {
                    "ok": True,
                    "action": "last_date_update",
                    "codigo": codigo_norm,
                    "before": before,
                    "after": after,
                    "no_change": True,
                    "history_preserved": True,
                }

            if cat is None and new_override is not None:
                con.execute(
                    """
                    INSERT INTO inventario_catalogo
                    (codigo, descripcion_estandar, unidad_base, categoria, tipo, ocurrencias, variaciones, activo, fuente, updated_at)
                    VALUES (?, '', '', '', '', 0, 'NO', 1, 'MANUAL_EDICION_CELDA', datetime('now'))
                    """,
                    (codigo_norm,),
                )

            if cat is not None or new_override is not None:
                con.execute(
                    """
                    UPDATE inventario_catalogo
                    SET ultima_fecha_override = ?,
                        updated_at = datetime('now')
                    WHERE codigo = ?
                    """,
                    (new_override if new_override else None, codigo_norm),
                )
                con.commit()

            return {
                "ok": True,
                "action": "last_date_update",
                "codigo": codigo_norm,
                "before": before,
                "after": after,
                "history_preserved": True,
            }

        # 4) Tipo de ultima modificacion (ENTRADA/SALIDA) como override visual
        if col == "ultima_modificacion_tipo":
            raw = _normalize_text(new_value).upper()
            type_map = {
                "ENTRADA": "ENTRADA",
                "E": "ENTRADA",
                "+": "ENTRADA",
                "SALIDA": "SALIDA",
                "S": "SALIDA",
                "-": "SALIDA",
            }
            if raw and raw not in type_map:
                return {"ok": False, "error": "Tipo invalido. Usa ENTRADA o SALIDA."}
            new_override = type_map.get(raw) if raw else None

            cat = con.execute(
                """
                SELECT codigo, ultima_tipo_override
                FROM inventario_catalogo
                WHERE codigo = ?
                """,
                (codigo_norm,),
            ).fetchone()
            last_mov = con.execute(
                """
                SELECT signo, tipo_mov
                FROM inventario_movimientos
                WHERE codigo = ?
                ORDER BY fecha DESC, id DESC
                LIMIT 1
                """,
                (codigo_norm,),
            ).fetchone()

            old_override_raw = _normalize_text(cat["ultima_tipo_override"]).upper() if cat else ""
            old_override = type_map.get(old_override_raw, old_override_raw or None)

            if last_mov is None:
                last_type = ""
            elif int(last_mov["signo"] or 0) == 1:
                last_type = "ENTRADA"
            elif int(last_mov["signo"] or 0) == -1:
                last_type = "SALIDA"
            else:
                last_type = _normalize_text(last_mov["tipo_mov"]).upper()

            before = old_override or last_type
            after = new_override or last_type

            if old_override == new_override:
                return {
                    "ok": True,
                    "action": "last_type_update",
                    "codigo": codigo_norm,
                    "before": before,
                    "after": after,
                    "no_change": True,
                    "history_preserved": True,
                }

            if cat is None and new_override is not None:
                con.execute(
                    """
                    INSERT INTO inventario_catalogo
                    (codigo, descripcion_estandar, unidad_base, categoria, tipo, ocurrencias, variaciones, activo, fuente, updated_at)
                    VALUES (?, '', '', '', '', 0, 'NO', 1, 'MANUAL_EDICION_CELDA', datetime('now'))
                    """,
                    (codigo_norm,),
                )

            if cat is not None or new_override is not None:
                con.execute(
                    """
                    UPDATE inventario_catalogo
                    SET ultima_tipo_override = ?,
                        updated_at = datetime('now')
                    WHERE codigo = ?
                    """,
                    (new_override if new_override else None, codigo_norm),
                )
                con.commit()

            return {
                "ok": True,
                "action": "last_type_update",
                "codigo": codigo_norm,
                "before": before,
                "after": after,
                "history_preserved": True,
            }

        return {"ok": False, "error": "Operacion no soportada."}
    finally:
        con.close()


def get_stock_summary(
    db_path: str,
    search_text: str = "",
    limit: int = 2000,
) -> List[Dict[str, Any]]:
    ensure_inventory_schema(db_path)
    con = _connect(db_path)
    try:
        text_value = _normalize_text(search_text)
        like_value = f"%{text_value}%"
        rows = con.execute(
            """
            WITH base AS (
                SELECT codigo FROM inventario_catalogo
                UNION
                SELECT DISTINCT codigo FROM inventario_movimientos
            ),
            last_mov AS (
                SELECT
                    m.codigo,
                    m.fecha,
                    m.tipo_mov,
                    m.signo,
                    ROW_NUMBER() OVER (
                        PARTITION BY m.codigo
                        ORDER BY m.fecha DESC, m.id DESC
                    ) AS rn
                FROM inventario_movimientos m
            )
            SELECT
                b.codigo AS codigo,
                COALESCE(c.descripcion_estandar, '') AS descripcion_estandar,
                COALESCE(c.unidad_base, '') AS unidad_base,
                COALESCE(c.categoria, '') AS categoria,
                COALESCE(c.tipo, '') AS tipo,
                COALESCE(SUM(CASE WHEN m.signo = 1 THEN m.cantidad ELSE 0 END), 0) AS entradas,
                COALESCE(SUM(CASE WHEN m.signo = -1 THEN m.cantidad ELSE 0 END), 0) AS salidas,
                COALESCE(SUM(m.signo * m.cantidad), 0) AS stock_actual,
                COALESCE(NULLIF(TRIM(COALESCE(c.ultima_fecha_override, '')), ''), COALESCE(lm.fecha, '')) AS ultima_fecha,
                CASE
                    WHEN NULLIF(TRIM(COALESCE(c.ultima_tipo_override, '')), '') IS NOT NULL
                        THEN UPPER(TRIM(COALESCE(c.ultima_tipo_override, '')))
                    WHEN lm.signo = 1 THEN 'ENTRADA'
                    WHEN lm.signo = -1 THEN 'SALIDA'
                    ELSE COALESCE(lm.tipo_mov, '')
                END AS ultima_modificacion_tipo
            FROM base b
            LEFT JOIN inventario_catalogo c
              ON c.codigo = b.codigo
            LEFT JOIN inventario_movimientos m
              ON m.codigo = b.codigo
            LEFT JOIN last_mov lm
              ON lm.codigo = b.codigo
             AND lm.rn = 1
            WHERE UPPER(TRIM(COALESCE(b.codigo, ''))) NOT IN ('', '-')
              AND UPPER(TRIM(COALESCE(c.categoria, ''))) NOT IN ('EPP', 'MATERIAL', 'HERRAMIENTA', 'CONTABLE', 'RIEGO')
              AND UPPER(TRIM(COALESCE(c.tipo, ''))) NOT IN (
                    'PROTECCION PERSONAL',
                    'INSUMO GENERAL',
                    'EQUIPO',
                    'NO INVENTARIABLE',
                    'INFRAESTRUCTURA'
              )
              AND (
                    (? = '')
                 OR b.codigo LIKE ?
                 OR COALESCE(c.descripcion_estandar, '') LIKE ?
              )
            GROUP BY
                b.codigo, c.descripcion_estandar, c.unidad_base, c.categoria, c.tipo,
                c.ultima_fecha_override, c.ultima_tipo_override,
                lm.fecha, lm.signo, lm.tipo_mov
            ORDER BY c.categoria, c.tipo, b.codigo
            LIMIT ?
            """,
            (text_value, like_value, like_value, max(1, int(limit))),
        ).fetchall()

        out: List[Dict[str, Any]] = []
        for r in rows:
            if not _is_field_product(r["categoria"], r["tipo"]):
                continue
            out.append(
                {
                    "codigo": r["codigo"],
                    "descripcion_estandar": r["descripcion_estandar"],
                    "unidad_base": r["unidad_base"],
                    "categoria": r["categoria"],
                    "tipo": r["tipo"],
                    "entradas": float(r["entradas"] or 0),
                    "salidas": float(r["salidas"] or 0),
                    "stock_actual": float(r["stock_actual"] or 0),
                    "ultima_fecha": r["ultima_fecha"] or "",
                    "ultima_modificacion_tipo": r["ultima_modificacion_tipo"] or "",
                }
            )
        return out
    finally:
        con.close()


def get_stock_for_code(db_path: str, codigo: str) -> float:
    ensure_inventory_schema(db_path)
    con = _connect(db_path)
    try:
        codigo_norm = _normalize_code(codigo)
        if not _is_valid_inventory_code(codigo_norm):
            return 0.0
        return _stock_for_code(con, codigo_norm)
    finally:
        con.close()


def _month_date_bounds(year: int, month: int) -> tuple[str, str]:
    y = int(year)
    m = int(month)
    if m < 1 or m > 12:
        raise ValueError("Mes invalido. Debe estar entre 1 y 12.")
    start = date(y, m, 1)
    next_month = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
    end = next_month - timedelta(days=1)
    return start.isoformat(), end.isoformat()


def _normalize_application_products(
    productos: Optional[List[Dict[str, Any]]],
) -> Dict[str, Any]:
    normalized: List[Dict[str, Any]] = []
    for idx, raw in enumerate(productos or [], start=1):
        if not isinstance(raw, dict):
            return {"ok": False, "error": f"Producto #{idx} invalido."}

        codigo = _normalize_code(raw.get("codigo"))
        if not _is_valid_inventory_code(codigo):
            return {"ok": False, "error": f"Codigo invalido en producto #{idx}."}

        cantidad = _to_float(raw.get("cantidad"))
        if cantidad <= 0:
            return {"ok": False, "error": f"Cantidad invalida en producto #{idx}."}

        normalized.append(
            {
                "codigo": codigo,
                "cantidad": cantidad,
                "unidad": _normalize_text(raw.get("unidad")).upper(),
                "observacion": _normalize_text(raw.get("observacion")),
            }
        )
    return {"ok": True, "products": normalized}


def create_field_application(
    db_path: str,
    titulo: str,
    fecha_programada: Any,
    descripcion: str = "",
    estado: str = "PROGRAMADA",
    productos: Optional[List[Dict[str, Any]]] = None,
    registrar_salidas: bool = False,
    fecha_ejecucion: Any = None,
    allow_negative: bool = False,
) -> Dict[str, Any]:
    ensure_inventory_schema(db_path)

    titulo_norm = _normalize_text(titulo)
    if not titulo_norm:
        return {"ok": False, "error": "Debes ingresar un titulo para la aplicacion."}

    fecha_prog = _parse_date_iso_strict(fecha_programada)
    if not fecha_prog:
        return {"ok": False, "error": "Fecha programada invalida. Usa YYYY-MM-DD."}

    estado_norm = _normalize_application_status(estado)
    if estado_norm not in _APP_STATUSES:
        return {"ok": False, "error": f"Estado invalido: {estado}"}

    normalized_products = _normalize_application_products(productos)
    if not normalized_products.get("ok"):
        return normalized_products

    fecha_ejec_iso: Optional[str] = None
    if estado_norm == "EJECUTADA":
        fecha_ejec_iso = _parse_date_iso_strict(fecha_ejecucion) if fecha_ejecucion else fecha_prog
        if not fecha_ejec_iso:
            return {"ok": False, "error": "Fecha de ejecucion invalida. Usa YYYY-MM-DD."}
    else:
        registrar_salidas = False

    con = _connect(db_path)
    try:
        cur = con.execute(
            """
            INSERT INTO aplicaciones_campo
            (titulo, descripcion, fecha_programada, fecha_ejecucion, estado, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, datetime('now'), datetime('now'))
            """,
            (
                titulo_norm,
                _normalize_text(descripcion),
                fecha_prog,
                fecha_ejec_iso,
                estado_norm,
            ),
        )
        app_id = int(cur.lastrowid or 0)
        movements_created = 0

        for p in normalized_products["products"]:
            movement_id = None
            if registrar_salidas:
                usage_obs = f"APLICACION_CAMPO #{app_id} - {titulo_norm}"
                if p["observacion"]:
                    usage_obs = f"{usage_obs} | {p['observacion']}"
                usage = _register_usage_in_connection(
                    con=con,
                    codigo=p["codigo"],
                    cantidad=p["cantidad"],
                    fecha_uso=fecha_ejec_iso or fecha_prog,
                    observacion=usage_obs,
                    unidad=p["unidad"],
                    allow_negative=allow_negative,
                    fuente="APLICACION_CAMPO",
                    referencia=f"APP:{app_id}",
                )
                if not usage.get("ok"):
                    con.rollback()
                    return {
                        "ok": False,
                        "error": usage.get("error", "No se pudo registrar salida de inventario."),
                        "codigo": usage.get("codigo", p["codigo"]),
                        "application_id": app_id,
                    }
                movement_id = int(usage.get("movement_id", 0) or 0) or None
                movements_created += 1

            con.execute(
                """
                INSERT INTO aplicaciones_campo_productos
                (aplicacion_id, codigo, cantidad, unidad, observacion, movimiento_id, created_at)
                VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
                """,
                (
                    app_id,
                    p["codigo"],
                    p["cantidad"],
                    p["unidad"],
                    p["observacion"],
                    movement_id,
                ),
            )

        con.commit()
        return {
            "ok": True,
            "application_id": app_id,
            "estado": estado_norm,
            "products_count": len(normalized_products["products"]),
            "movements_created": movements_created,
            "fecha_programada": fecha_prog,
            "fecha_ejecucion": fecha_ejec_iso,
        }
    finally:
        con.close()


def list_field_applications(
    db_path: str,
    date_from: Any = None,
    date_to: Any = None,
    status: str = "",
    search_text: str = "",
    limit: int = 2000,
) -> List[Dict[str, Any]]:
    ensure_inventory_schema(db_path)
    status_norm = _normalize_application_status(status) if _normalize_text(status) else ""
    if status_norm and status_norm not in _APP_STATUSES:
        status_norm = ""

    date_from_iso = _parse_date_iso_strict(date_from) if _normalize_text(date_from) else ""
    date_to_iso = _parse_date_iso_strict(date_to) if _normalize_text(date_to) else ""
    search_norm = _normalize_text(search_text).upper()
    like_value = f"%{search_norm}%"

    con = _connect(db_path)
    try:
        rows = con.execute(
            """
            SELECT
                a.id,
                a.titulo,
                COALESCE(a.descripcion, '') AS descripcion,
                a.fecha_programada,
                COALESCE(a.fecha_ejecucion, '') AS fecha_ejecucion,
                UPPER(TRIM(COALESCE(a.estado, 'PROGRAMADA'))) AS estado,
                COALESCE(COUNT(p.id), 0) AS productos_total,
                COALESCE(SUM(CASE WHEN p.movimiento_id IS NOT NULL THEN 1 ELSE 0 END), 0) AS productos_con_salida
            FROM aplicaciones_campo a
            LEFT JOIN aplicaciones_campo_productos p
              ON p.aplicacion_id = a.id
            WHERE (? = '' OR UPPER(TRIM(COALESCE(a.estado, 'PROGRAMADA'))) = ?)
              AND (? = '' OR a.fecha_programada >= ?)
              AND (? = '' OR a.fecha_programada <= ?)
              AND (
                    (? = '')
                 OR UPPER(COALESCE(a.titulo, '')) LIKE ?
                 OR UPPER(COALESCE(a.descripcion, '')) LIKE ?
              )
            GROUP BY
                a.id, a.titulo, a.descripcion, a.fecha_programada, a.fecha_ejecucion, a.estado
            ORDER BY a.fecha_programada ASC, a.id ASC
            LIMIT ?
            """,
            (
                status_norm,
                status_norm,
                date_from_iso or "",
                date_from_iso or "",
                date_to_iso or "",
                date_to_iso or "",
                search_norm,
                like_value,
                like_value,
                max(1, int(limit)),
            ),
        ).fetchall()
        return [
            {
                "id": int(r["id"]),
                "titulo": r["titulo"] or "",
                "descripcion": r["descripcion"] or "",
                "fecha_programada": r["fecha_programada"] or "",
                "fecha_ejecucion": r["fecha_ejecucion"] or "",
                "estado": (r["estado"] or "PROGRAMADA").upper(),
                "productos_total": int(r["productos_total"] or 0),
                "productos_con_salida": int(r["productos_con_salida"] or 0),
            }
            for r in rows
        ]
    finally:
        con.close()


def get_field_application(db_path: str, application_id: int) -> Dict[str, Any]:
    ensure_inventory_schema(db_path)
    app_id = int(application_id or 0)
    if app_id <= 0:
        return {"ok": False, "error": "ID de aplicacion invalido."}

    con = _connect(db_path)
    try:
        row = con.execute(
            """
            SELECT
                id,
                titulo,
                COALESCE(descripcion, '') AS descripcion,
                fecha_programada,
                COALESCE(fecha_ejecucion, '') AS fecha_ejecucion,
                UPPER(TRIM(COALESCE(estado, 'PROGRAMADA'))) AS estado
            FROM aplicaciones_campo
            WHERE id = ?
            """,
            (app_id,),
        ).fetchone()
        if row is None:
            return {"ok": False, "error": f"No existe la aplicacion #{app_id}."}

        products = con.execute(
            """
            SELECT
                p.id,
                p.codigo,
                COALESCE(c.descripcion_estandar, '') AS descripcion_estandar,
                p.cantidad,
                COALESCE(p.unidad, '') AS unidad,
                COALESCE(p.observacion, '') AS observacion,
                p.movimiento_id
            FROM aplicaciones_campo_productos p
            LEFT JOIN inventario_catalogo c
              ON c.codigo = p.codigo
            WHERE p.aplicacion_id = ?
            ORDER BY p.id ASC
            """,
            (app_id,),
        ).fetchall()

        return {
            "ok": True,
            "application": {
                "id": int(row["id"]),
                "titulo": row["titulo"] or "",
                "descripcion": row["descripcion"] or "",
                "fecha_programada": row["fecha_programada"] or "",
                "fecha_ejecucion": row["fecha_ejecucion"] or "",
                "estado": (row["estado"] or "PROGRAMADA").upper(),
            },
            "products": [
                {
                    "id": int(p["id"]),
                    "codigo": p["codigo"] or "",
                    "descripcion_estandar": p["descripcion_estandar"] or "",
                    "cantidad": float(p["cantidad"] or 0),
                    "unidad": p["unidad"] or "",
                    "observacion": p["observacion"] or "",
                    "movimiento_id": int(p["movimiento_id"] or 0) if p["movimiento_id"] is not None else None,
                }
                for p in products
            ],
        }
    finally:
        con.close()


def list_field_application_products(db_path: str, application_id: int) -> List[Dict[str, Any]]:
    ensure_inventory_schema(db_path)
    app_id = int(application_id or 0)
    if app_id <= 0:
        return []

    con = _connect(db_path)
    try:
        rows = con.execute(
            """
            SELECT
                p.id,
                p.aplicacion_id,
                p.codigo,
                COALESCE(c.descripcion_estandar, '') AS descripcion_estandar,
                p.cantidad,
                COALESCE(p.unidad, '') AS unidad,
                COALESCE(p.observacion, '') AS observacion,
                p.movimiento_id,
                COALESCE(m.fecha, '') AS fecha_movimiento
            FROM aplicaciones_campo_productos p
            LEFT JOIN inventario_catalogo c
              ON c.codigo = p.codigo
            LEFT JOIN inventario_movimientos m
              ON m.id = p.movimiento_id
            WHERE p.aplicacion_id = ?
            ORDER BY p.id ASC
            """,
            (app_id,),
        ).fetchall()
        return [
            {
                "id": int(r["id"]),
                "aplicacion_id": int(r["aplicacion_id"]),
                "codigo": r["codigo"] or "",
                "descripcion_estandar": r["descripcion_estandar"] or "",
                "cantidad": float(r["cantidad"] or 0),
                "unidad": r["unidad"] or "",
                "observacion": r["observacion"] or "",
                "movimiento_id": int(r["movimiento_id"] or 0) if r["movimiento_id"] is not None else None,
                "fecha_movimiento": r["fecha_movimiento"] or "",
            }
            for r in rows
        ]
    finally:
        con.close()


def update_field_application(
    db_path: str,
    application_id: int,
    titulo: str,
    fecha_programada: Any,
    descripcion: str = "",
    productos: Optional[List[Dict[str, Any]]] = None,
    allow_negative: bool = False,
) -> Dict[str, Any]:
    """
    Permite editar una aplicacion en su totalidad:
      - titulo
      - fecha_programada
      - descripcion
      - productos

    Reglas:
      - Si esta PROGRAMADA: se actualiza el plan (sin movimientos de inventario).
      - Si esta EJECUTADA: se corrige la ejecucion real.
        Se eliminan los movimientos anteriores de esa aplicacion y se recrean
        segun el nuevo detalle de productos.
    """
    ensure_inventory_schema(db_path)
    app_id = int(application_id or 0)
    if app_id <= 0:
        return {"ok": False, "error": "ID de aplicacion invalido."}

    titulo_norm = _normalize_text(titulo)
    if not titulo_norm:
        return {"ok": False, "error": "Debes ingresar un titulo para la aplicacion."}

    fecha_prog = _parse_date_iso_strict(fecha_programada)
    if not fecha_prog:
        return {"ok": False, "error": "Fecha programada invalida. Usa YYYY-MM-DD."}

    normalized_products = _normalize_application_products(productos)
    if not normalized_products.get("ok"):
        return normalized_products

    con = _connect(db_path)
    try:
        app = con.execute(
            """
            SELECT id, estado, fecha_ejecucion
            FROM aplicaciones_campo
            WHERE id = ?
            """,
            (app_id,),
        ).fetchone()
        if app is None:
            return {"ok": False, "error": f"No existe la aplicacion #{app_id}."}

        estado_actual = _normalize_application_status(app["estado"])
        if estado_actual not in {"PROGRAMADA", "EJECUTADA"}:
            return {"ok": False, "error": f"No se puede editar estado actual: {estado_actual}"}

        existing_products = con.execute(
            """
            SELECT movimiento_id
            FROM aplicaciones_campo_productos
            WHERE aplicacion_id = ?
            """,
            (app_id,),
        ).fetchall()

        movement_ids: list[int] = []
        for r in existing_products:
            mov = r["movimiento_id"]
            if mov is None:
                continue
            try:
                movement_ids.append(int(mov))
            except Exception:
                pass

        con.execute(
            """
            UPDATE aplicaciones_campo
            SET titulo = ?,
                descripcion = ?,
                fecha_programada = ?,
                updated_at = datetime('now')
            WHERE id = ?
            """,
            (
                titulo_norm,
                _normalize_text(descripcion),
                fecha_prog,
                app_id,
            ),
        )

        con.execute(
            "DELETE FROM aplicaciones_campo_productos WHERE aplicacion_id = ?",
            (app_id,),
        )

        movements_deleted = 0
        if movement_ids:
            placeholders = ", ".join("?" for _ in movement_ids)
            movements_deleted += int(
                con.execute(
                    f"DELETE FROM inventario_movimientos WHERE id IN ({placeholders})",
                    tuple(movement_ids),
                ).rowcount
                or 0
            )

        # Limpieza defensiva para movimientos huérfanos de esta aplicación.
        movements_deleted += int(
            con.execute(
                """
                DELETE FROM inventario_movimientos
                WHERE UPPER(TRIM(COALESCE(fuente, ''))) = 'APLICACION_CAMPO'
                  AND UPPER(TRIM(COALESCE(referencia, ''))) = ?
                """,
                (f"APP:{app_id}".upper(),),
            ).rowcount
            or 0
        )

        movements_created = 0
        if estado_actual == "EJECUTADA":
            fecha_ejec = _parse_date_iso_strict(app["fecha_ejecucion"]) if app["fecha_ejecucion"] else None
            fecha_ejec = fecha_ejec or fecha_prog

            for p in normalized_products["products"]:
                usage_obs = f"APLICACION_CAMPO #{app_id} - {titulo_norm}"
                if p["observacion"]:
                    usage_obs = f"{usage_obs} | {p['observacion']}"

                usage = _register_usage_in_connection(
                    con=con,
                    codigo=p["codigo"],
                    cantidad=p["cantidad"],
                    fecha_uso=fecha_ejec,
                    observacion=usage_obs,
                    unidad=p["unidad"],
                    allow_negative=allow_negative,
                    fuente="APLICACION_CAMPO",
                    referencia=f"APP:{app_id}",
                )
                if not usage.get("ok"):
                    con.rollback()
                    return {
                        "ok": False,
                        "error": usage.get("error", "No se pudo corregir salidas para la aplicación."),
                        "codigo": usage.get("codigo", p["codigo"]),
                        "application_id": app_id,
                    }

                movement_id = int(usage.get("movement_id", 0) or 0)
                con.execute(
                    """
                    INSERT INTO aplicaciones_campo_productos
                    (aplicacion_id, codigo, cantidad, unidad, observacion, movimiento_id, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
                    """,
                    (
                        app_id,
                        p["codigo"],
                        p["cantidad"],
                        p["unidad"],
                        p["observacion"],
                        movement_id if movement_id > 0 else None,
                    ),
                )
                movements_created += 1

            con.execute(
                """
                UPDATE aplicaciones_campo
                SET estado = 'EJECUTADA',
                    fecha_ejecucion = ?,
                    updated_at = datetime('now')
                WHERE id = ?
                """,
                (fecha_ejec, app_id),
            )
        else:
            for p in normalized_products["products"]:
                con.execute(
                    """
                    INSERT INTO aplicaciones_campo_productos
                    (aplicacion_id, codigo, cantidad, unidad, observacion, movimiento_id, created_at)
                    VALUES (?, ?, ?, ?, ?, NULL, datetime('now'))
                    """,
                    (
                        app_id,
                        p["codigo"],
                        p["cantidad"],
                        p["unidad"],
                        p["observacion"],
                    ),
                )

            con.execute(
                """
                UPDATE aplicaciones_campo
                SET estado = 'PROGRAMADA',
                    fecha_ejecucion = NULL,
                    updated_at = datetime('now')
                WHERE id = ?
                """,
                (app_id,),
            )

        con.commit()
        return {
            "ok": True,
            "application_id": app_id,
            "estado": "EJECUTADA" if estado_actual == "EJECUTADA" else "PROGRAMADA",
            "products_count": len(normalized_products["products"]),
            "fecha_programada": fecha_prog,
            "movements_deleted": movements_deleted,
            "movements_created": movements_created,
        }
    finally:
        con.close()


def delete_field_application(
    db_path: str,
    application_id: int,
) -> Dict[str, Any]:
    """
    Elimina una aplicacion y sus salidas de inventario asociadas.
    """
    ensure_inventory_schema(db_path)
    app_id = int(application_id or 0)
    if app_id <= 0:
        return {"ok": False, "error": "ID de aplicacion invalido."}

    con = _connect(db_path)
    try:
        app = con.execute(
            """
            SELECT id, estado
            FROM aplicaciones_campo
            WHERE id = ?
            """,
            (app_id,),
        ).fetchone()
        if app is None:
            return {"ok": False, "error": f"No existe la aplicacion #{app_id}."}

        rows_mov = con.execute(
            """
            SELECT DISTINCT movimiento_id
            FROM aplicaciones_campo_productos
            WHERE aplicacion_id = ?
              AND movimiento_id IS NOT NULL
            """,
            (app_id,),
        ).fetchall()
        movement_ids: list[int] = []
        for r in rows_mov:
            mov = r["movimiento_id"]
            if mov is None:
                continue
            try:
                movement_ids.append(int(mov))
            except Exception:
                pass

        n_products = con.execute(
            "SELECT COUNT(*) AS n FROM aplicaciones_campo_productos WHERE aplicacion_id = ?",
            (app_id,),
        ).fetchone()
        products_count = int(n_products["n"] or 0) if n_products else 0

        con.execute("DELETE FROM aplicaciones_campo_productos WHERE aplicacion_id = ?", (app_id,))
        con.execute("DELETE FROM aplicaciones_campo WHERE id = ?", (app_id,))

        movements_deleted = 0
        if movement_ids:
            placeholders = ", ".join("?" for _ in movement_ids)
            movements_deleted += int(
                con.execute(
                    f"DELETE FROM inventario_movimientos WHERE id IN ({placeholders})",
                    tuple(movement_ids),
                ).rowcount
                or 0
            )

        movements_deleted += int(
            con.execute(
                """
                DELETE FROM inventario_movimientos
                WHERE UPPER(TRIM(COALESCE(fuente, ''))) = 'APLICACION_CAMPO'
                  AND UPPER(TRIM(COALESCE(referencia, ''))) = ?
                """,
                (f"APP:{app_id}".upper(),),
            ).rowcount
            or 0
        )

        con.commit()
        return {
            "ok": True,
            "application_id": app_id,
            "products_deleted": products_count,
            "estado_anterior": _normalize_application_status(app["estado"]),
            "movements_deleted": movements_deleted,
        }
    finally:
        con.close()


def replace_field_application_products(
    db_path: str,
    application_id: int,
    productos: Optional[List[Dict[str, Any]]],
    require_no_movements: bool = True,
) -> Dict[str, Any]:
    """
    Reemplaza el detalle de productos de una aplicacion.

    Uso principal:
      - Guardar productos reales al momento de ejecutar una aplicacion.
    """
    ensure_inventory_schema(db_path)
    app_id = int(application_id or 0)
    if app_id <= 0:
        return {"ok": False, "error": "ID de aplicacion invalido."}

    normalized_products = _normalize_application_products(productos)
    if not normalized_products.get("ok"):
        return normalized_products

    products = normalized_products["products"]
    if not products:
        return {"ok": False, "error": "Debes ingresar al menos un producto real."}

    con = _connect(db_path)
    try:
        app = con.execute(
            "SELECT id FROM aplicaciones_campo WHERE id = ?",
            (app_id,),
        ).fetchone()
        if app is None:
            return {"ok": False, "error": f"No existe la aplicacion #{app_id}."}

        if require_no_movements:
            row = con.execute(
                """
                SELECT COUNT(*) AS n
                FROM aplicaciones_campo_productos
                WHERE aplicacion_id = ?
                  AND movimiento_id IS NOT NULL
                """,
                (app_id,),
            ).fetchone()
            linked = int(row["n"] or 0) if row else 0
            if linked > 0:
                return {
                    "ok": False,
                    "error": "La aplicacion ya tiene salidas registradas y no se puede reemplazar el detalle real.",
                    "application_id": app_id,
                    "linked_movements": linked,
                }

        con.execute(
            "DELETE FROM aplicaciones_campo_productos WHERE aplicacion_id = ?",
            (app_id,),
        )
        for p in products:
            con.execute(
                """
                INSERT INTO aplicaciones_campo_productos
                (aplicacion_id, codigo, cantidad, unidad, observacion, movimiento_id, created_at)
                VALUES (?, ?, ?, ?, ?, NULL, datetime('now'))
                """,
                (
                    app_id,
                    p["codigo"],
                    p["cantidad"],
                    p["unidad"],
                    p["observacion"],
                ),
            )

        con.execute(
            """
            UPDATE aplicaciones_campo
            SET updated_at = datetime('now')
            WHERE id = ?
            """,
            (app_id,),
        )
        con.commit()
        return {
            "ok": True,
            "application_id": app_id,
            "products_count": len(products),
        }
    finally:
        con.close()


def execute_field_application(
    db_path: str,
    application_id: int,
    fecha_ejecucion: Any = None,
    registrar_salidas: bool = True,
    allow_negative: bool = False,
) -> Dict[str, Any]:
    ensure_inventory_schema(db_path)
    app_id = int(application_id or 0)
    if app_id <= 0:
        return {"ok": False, "error": "ID de aplicacion invalido."}

    fecha_exec = _parse_date_iso_strict(fecha_ejecucion) if fecha_ejecucion else date.today().isoformat()
    if not fecha_exec:
        return {"ok": False, "error": "Fecha de ejecucion invalida. Usa YYYY-MM-DD."}

    con = _connect(db_path)
    try:
        app = con.execute(
            """
            SELECT id, titulo, fecha_programada, estado
            FROM aplicaciones_campo
            WHERE id = ?
            """,
            (app_id,),
        ).fetchone()
        if app is None:
            return {"ok": False, "error": f"No existe la aplicacion #{app_id}."}

        products = con.execute(
            """
            SELECT id, codigo, cantidad, unidad, observacion, movimiento_id
            FROM aplicaciones_campo_productos
            WHERE aplicacion_id = ?
            ORDER BY id ASC
            """,
            (app_id,),
        ).fetchall()

        total_products = len(products)
        pending_before = sum(1 for p in products if p["movimiento_id"] is None)
        movements_created = 0

        if registrar_salidas:
            for p in products:
                if p["movimiento_id"] is not None:
                    continue
                usage_obs = f"APLICACION_CAMPO #{app_id} - {app['titulo'] or ''}".strip(" -")
                extra_obs = _normalize_text(p["observacion"])
                if extra_obs:
                    usage_obs = f"{usage_obs} | {extra_obs}" if usage_obs else extra_obs

                usage = _register_usage_in_connection(
                    con=con,
                    codigo=p["codigo"],
                    cantidad=float(p["cantidad"] or 0),
                    fecha_uso=fecha_exec,
                    observacion=usage_obs,
                    unidad=p["unidad"] or "",
                    allow_negative=allow_negative,
                    fuente="APLICACION_CAMPO",
                    referencia=f"APP:{app_id}",
                )
                if not usage.get("ok"):
                    con.rollback()
                    return {
                        "ok": False,
                        "error": usage.get("error", "No se pudo registrar salidas para la aplicacion."),
                        "codigo": usage.get("codigo", p["codigo"]),
                        "application_id": app_id,
                    }

                con.execute(
                    """
                    UPDATE aplicaciones_campo_productos
                    SET movimiento_id = ?
                    WHERE id = ?
                    """,
                    (int(usage.get("movement_id", 0) or 0), int(p["id"])),
                )
                movements_created += 1

        con.execute(
            """
            UPDATE aplicaciones_campo
            SET estado = 'EJECUTADA',
                fecha_ejecucion = ?,
                updated_at = datetime('now')
            WHERE id = ?
            """,
            (fecha_exec, app_id),
        )
        con.commit()
        return {
            "ok": True,
            "application_id": app_id,
            "fecha_ejecucion": fecha_exec,
            "estado": "EJECUTADA",
            "total_products": total_products,
            "pending_before": pending_before,
            "movements_created": movements_created,
        }
    finally:
        con.close()


def update_field_application_status(
    db_path: str,
    application_id: int,
    estado: str,
) -> Dict[str, Any]:
    ensure_inventory_schema(db_path)
    app_id = int(application_id or 0)
    if app_id <= 0:
        return {"ok": False, "error": "ID de aplicacion invalido."}

    status_norm = _normalize_application_status(estado)
    if status_norm not in _APP_STATUSES:
        return {"ok": False, "error": f"Estado invalido: {estado}"}

    if status_norm == "CANCELADA":
        return delete_field_application(
            db_path=db_path,
            application_id=app_id,
        )

    if status_norm == "EJECUTADA":
        return execute_field_application(
            db_path=db_path,
            application_id=app_id,
            fecha_ejecucion=date.today().isoformat(),
            registrar_salidas=False,
        )

    con = _connect(db_path)
    try:
        row = con.execute(
            "SELECT id FROM aplicaciones_campo WHERE id = ?",
            (app_id,),
        ).fetchone()
        if row is None:
            return {"ok": False, "error": f"No existe la aplicacion #{app_id}."}

        con.execute(
            """
            UPDATE aplicaciones_campo
            SET estado = ?,
                updated_at = datetime('now')
            WHERE id = ?
            """,
            (status_norm, app_id),
        )
        con.commit()
        return {"ok": True, "application_id": app_id, "estado": status_norm}
    finally:
        con.close()


def get_field_calendar_summary(
    db_path: str,
    year: int,
    month: int,
    status: str = "",
) -> Dict[str, Dict[str, int]]:
    ensure_inventory_schema(db_path)
    start_date, end_date = _month_date_bounds(int(year), int(month))
    status_norm = _normalize_application_status(status) if _normalize_text(status) else ""
    if status_norm and status_norm not in _APP_STATUSES:
        status_norm = ""

    con = _connect(db_path)
    try:
        rows = con.execute(
            """
            SELECT
                fecha_programada,
                COUNT(*) AS total,
                SUM(CASE WHEN UPPER(TRIM(COALESCE(estado, 'PROGRAMADA'))) = 'PROGRAMADA' THEN 1 ELSE 0 END) AS programadas,
                SUM(CASE WHEN UPPER(TRIM(COALESCE(estado, 'PROGRAMADA'))) = 'EJECUTADA' THEN 1 ELSE 0 END) AS ejecutadas,
                SUM(CASE WHEN UPPER(TRIM(COALESCE(estado, 'PROGRAMADA'))) = 'CANCELADA' THEN 1 ELSE 0 END) AS canceladas
            FROM aplicaciones_campo
            WHERE fecha_programada >= ?
              AND fecha_programada <= ?
              AND (? = '' OR UPPER(TRIM(COALESCE(estado, 'PROGRAMADA'))) = ?)
            GROUP BY fecha_programada
            ORDER BY fecha_programada ASC
            """,
            (start_date, end_date, status_norm, status_norm),
        ).fetchall()
        out: Dict[str, Dict[str, int]] = {}
        for r in rows:
            key = _normalize_text(r["fecha_programada"])
            if not key:
                continue
            out[key] = {
                "total": int(r["total"] or 0),
                "programadas": int(r["programadas"] or 0),
                "ejecutadas": int(r["ejecutadas"] or 0),
                "canceladas": int(r["canceladas"] or 0),
            }
        return out
    finally:
        con.close()


def _get_purchase_entries_for_export(db_path: str) -> List[Dict[str, Any]]:
    ensure_inventory_schema(db_path)
    con = _connect(db_path)
    try:
        has_detalle = _table_exists(con, "detalle")
        has_documentos = _table_exists(con, "documentos")
        detalle_cols = (
            {r[1] for r in con.execute("PRAGMA table_info(detalle);").fetchall()}
            if has_detalle
            else set()
        )
        documentos_cols = (
            {r[1] for r in con.execute("PRAGMA table_info(documentos);").fetchall()}
            if has_documentos
            else set()
        )

        select_parts = [
            "m.id AS movimiento_id",
            "COALESCE(m.fecha, '') AS fecha",
            "COALESCE(m.codigo, '') AS codigo",
            "COALESCE(c.descripcion_estandar, '') AS descripcion_estandar_catalogo",
            "COALESCE(c.categoria, '') AS categoria",
            "COALESCE(c.tipo, '') AS tipo",
            "COALESCE(m.cantidad, 0) AS cantidad_entrada",
            "COALESCE(m.unidad, '') AS unidad_movimiento",
            "COALESCE(m.referencia, '') AS referencia_detalle",
        ]

        join_parts = [
            "LEFT JOIN inventario_catalogo c ON c.codigo = m.codigo",
        ]

        if has_detalle:
            join_parts.append("LEFT JOIN detalle d ON d.id_det = m.referencia")
            select_parts.extend(
                [
                    "COALESCE(d.id_doc, '') AS id_doc" if "id_doc" in detalle_cols else "'' AS id_doc",
                    "COALESCE(d.linea, '') AS linea" if "linea" in detalle_cols else "'' AS linea",
                    "COALESCE(d.descripcion, '') AS descripcion_dte_original" if "descripcion" in detalle_cols else "'' AS descripcion_dte_original",
                    "COALESCE(d.cantidad, 0) AS cantidad_dte" if "cantidad" in detalle_cols else "0 AS cantidad_dte",
                    "COALESCE(d.unidad, '') AS unidad_dte" if "unidad" in detalle_cols else "'' AS unidad_dte",
                    "COALESCE(d.precio_unitario, 0) AS precio_unitario_dte" if "precio_unitario" in detalle_cols else "0 AS precio_unitario_dte",
                    "COALESCE(d.monto_item, 0) AS monto_item_dte" if "monto_item" in detalle_cols else "0 AS monto_item_dte",
                ]
            )
        else:
            select_parts.extend(
                [
                    "'' AS id_doc",
                    "'' AS linea",
                    "'' AS descripcion_dte_original",
                    "0 AS cantidad_dte",
                    "'' AS unidad_dte",
                    "0 AS precio_unitario_dte",
                    "0 AS monto_item_dte",
                ]
            )

        can_join_documentos = has_detalle and has_documentos and "id_doc" in detalle_cols
        if can_join_documentos:
            join_parts.append("LEFT JOIN documentos doc ON doc.id_doc = d.id_doc")
            select_parts.extend(
                [
                    "COALESCE(doc.razon_social, '') AS proveedor" if "razon_social" in documentos_cols else "'' AS proveedor",
                    "COALESCE(doc.rut_emisor, '') AS rut_emisor" if "rut_emisor" in documentos_cols else "'' AS rut_emisor",
                ]
            )
        else:
            select_parts.extend(
                [
                    "'' AS proveedor",
                    "'' AS rut_emisor",
                ]
            )

        query = f"""
            SELECT
                {", ".join(select_parts)}
            FROM inventario_movimientos m
            {' '.join(join_parts)}
            WHERE m.signo = 1
              AND COALESCE(m.fuente, '') = 'DTE_DETALLE'
              AND UPPER(TRIM(COALESCE(m.codigo, ''))) NOT IN ('', '-')
            ORDER BY m.fecha DESC, m.id DESC
        """

        rows = con.execute(query).fetchall()
        out: List[Dict[str, Any]] = []
        for r in rows:
            descripcion_catalogo = _normalize_text(r["descripcion_estandar_catalogo"])
            descripcion_dte = _normalize_text(r["descripcion_dte_original"])
            descripcion_final = descripcion_catalogo or descripcion_dte

            cant_dte = _to_float(r["cantidad_dte"])
            precio_unitario = _to_float(r["precio_unitario_dte"])
            monto_item = _to_float(r["monto_item_dte"])
            precio_calculado = round(monto_item / cant_dte, 6) if cant_dte > 0 else precio_unitario

            out.append(
                {
                    "movimiento_id": int(r["movimiento_id"] or 0),
                    "fecha": _normalize_text(r["fecha"]),
                    "codigo": _normalize_code(r["codigo"]),
                    "descripcion_estandar": descripcion_final,
                    "categoria": _normalize_text(r["categoria"]).upper(),
                    "tipo": _normalize_text(r["tipo"]).upper(),
                    "cantidad_entrada": _to_float(r["cantidad_entrada"]),
                    "unidad_movimiento": _normalize_text(r["unidad_movimiento"]).upper(),
                    "id_doc": _normalize_text(r["id_doc"]),
                    "linea": _normalize_text(r["linea"]),
                    "proveedor": _normalize_text(r["proveedor"]),
                    "rut_emisor": _normalize_text(r["rut_emisor"]),
                    "descripcion_dte_original": descripcion_dte,
                    "cantidad_dte": cant_dte,
                    "unidad_dte": _normalize_text(r["unidad_dte"]).upper(),
                    "precio_unitario_dte": precio_unitario,
                    "monto_item_dte": monto_item,
                    "precio_unitario_calculado": precio_calculado,
                    "referencia_detalle": _normalize_text(r["referencia_detalle"]),
                }
            )
        return out
    finally:
        con.close()


def export_stock_to_excel(db_path: str, output_path: str) -> Dict[str, Any]:
    try:
        from openpyxl import Workbook
    except Exception:
        return {"ok": False, "error": "Falta la dependencia openpyxl para exportar Excel."}

    rows = get_stock_summary(db_path=db_path, search_text="", limit=200000)
    entries_rows = _get_purchase_entries_for_export(db_path=db_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "stock"

    headers = [
        "codigo",
        "descripcion_estandar",
        "unidad_base",
        "categoria",
        "tipo",
        "entradas",
        "salidas",
        "stock_actual",
        "ultima_fecha",
        "ultima_modificacion_tipo",
    ]
    ws.append(headers)

    for r in rows:
        ws.append(
            [
                r["codigo"],
                r["descripcion_estandar"],
                r["unidad_base"],
                r["categoria"],
                r["tipo"],
                r["entradas"],
                r["salidas"],
                r["stock_actual"],
                r["ultima_fecha"],
                r["ultima_modificacion_tipo"],
            ]
        )

    ws_entries = wb.create_sheet(title="entradas_compra")
    entries_headers = [
        "movimiento_id",
        "fecha",
        "codigo",
        "descripcion_estandar",
        "categoria",
        "tipo",
        "cantidad_entrada",
        "unidad_movimiento",
        "id_doc",
        "linea",
        "proveedor",
        "rut_emisor",
        "descripcion_dte_original",
        "cantidad_dte",
        "unidad_dte",
        "precio_unitario_dte",
        "monto_item_dte",
        "precio_unitario_calculado",
        "referencia_detalle",
    ]
    ws_entries.append(entries_headers)

    for r in entries_rows:
        ws_entries.append(
            [
                r["movimiento_id"],
                r["fecha"],
                r["codigo"],
                r["descripcion_estandar"],
                r["categoria"],
                r["tipo"],
                r["cantidad_entrada"],
                r["unidad_movimiento"],
                r["id_doc"],
                r["linea"],
                r["proveedor"],
                r["rut_emisor"],
                r["descripcion_dte_original"],
                r["cantidad_dte"],
                r["unidad_dte"],
                r["precio_unitario_dte"],
                r["monto_item_dte"],
                r["precio_unitario_calculado"],
                r["referencia_detalle"],
            ]
        )

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))

    return {
        "ok": True,
        "path": str(out),
        "n_rows": len(rows),
        "n_entries_rows": len(entries_rows),
    }
